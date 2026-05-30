#!/usr/bin/env python3
"""
Opportunity Hunter V6 - UNIFIED EDITION
==========================================
Merges OH2 V5 (original) + NO-CAPTCHA Hunter (v2)
+ Captcha detection + Site type filtering + Scrapy crawling

Runs daily at 00:00 UTC on GitHub Actions.
Outputs: Google Sheet (daily tab), Google Doc, top_finds.json
"""

import os, json, time, re, hashlib, sys, urllib.parse, html as html_mod, random, smtplib, base64
from datetime import datetime, timezone
from typing import Optional
from dataclasses import dataclass, field
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

os.system("pip install requests openpyxl google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client playwright google-generativeai curl-cffi 2>/dev/null")
os.system("playwright install chromium 2>/dev/null")

import google.generativeai as genai
import requests
from openpyxl import Workbook, load_workbook

# === API KEYS ===

FIRECRAWL_API_KEY = os.environ.get("FIRECRAWL_API_KEY", "")
EXA_API_KEY = os.environ.get("EXA_API_KEY", "")
GEMINI_API_KEYS = []
_key = os.environ.get("GEMINI_API_KEY", "")
if _key: GEMINI_API_KEYS.append(_key)
for _i in range(2, 10):
    _key = os.environ.get(f"GEMINI_API_KEY_{_i}", "")
    if _key: GEMINI_API_KEYS.append(_key)
GROQ_API_KEYS = []
for _i in range(1, 10):
    _key = os.environ.get(f"GROQ_API_KEY_{_i}" if _i > 1 else "GROQ_API_KEY", "")
    if _key: GROQ_API_KEYS.append(_key)
GOOGLE_OAUTH_CLIENT_ID = os.environ.get("GOOGLE_OAUTH_CLIENT_ID", "")
GOOGLE_OAUTH_CLIENT_SECRET = os.environ.get("GOOGLE_OAUTH_CLIENT_SECRET", "")
GOOGLE_REFRESH_TOKEN = os.environ.get("GOOGLE_REFRESH_TOKEN", "")
GOOGLE_SERVICE_ACCOUNT_JSON = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
SERVICE_ACCOUNT_INFO = None
if GOOGLE_SERVICE_ACCOUNT_JSON:
    try:
        SERVICE_ACCOUNT_INFO = json.loads(GOOGLE_SERVICE_ACCOUNT_JSON)
    except (json.JSONDecodeError, TypeError):
        pass
OPENROUTER_API_KEYS = []
_key = os.environ.get("OPENROUTER_API_KEY", "")
if _key: OPENROUTER_API_KEYS.append(_key)
for _i in range(2, 10):
    _key = os.environ.get(f"OPENROUTER_API_KEY_{_i}", "")
    if _key: OPENROUTER_API_KEYS.append(_key)

GMAIL_USER = os.environ.get("GMAIL_USER", "wo312092@gmail.com")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
NOTIFICATION_EMAIL = os.environ.get("NOTIFICATION_EMAIL", "wo312092@gmail.com")
GOOGLE_DOC_ID = "1tuZYC0RoCxP-Js5FuSUy6BmA72xU57N0M80O-ksoXoE"
HARDCODED_SHEET_ID = "16N4vLfARruxPC4SqUv6tx5sWJi35ay-5hIu5UyQ84XE"
EXCEL_FILE = "opportunities.xlsx"
MEMORY_FILE = "bot_memory.json"
REPORT_DIR = "reports"
TOP_FINDS_FILE = "top_automatable_finds.json"
DEEP_ANALYSIS_DIR = "deep_analysis"

_gemini_key_index = 0
_groq_key_index = 0
_openrouter_key_index = 0

DEFAULT_GROQ_MODEL = "llama-3.3-70b-versatile"
FAST_GROQ_MODEL = "llama-3.1-8b-instant"

# === ALL QUERIES (MERGED + DEDUPLICATED) ===

ALL_QUERIES = [
    # -- BITCOIN --
    "free bitcoin faucet instant withdrawal 2026 no survey",
    "btc faucet auto claim free bitcoin 2026",
    "bitcoin mining free cloud mining 2026 no deposit withdraw",
    "earn free satoshi every hour 2026 btc faucet",
    "bitcoin faucet instant pay to wallet 2026",
    "btc mining telegram bot free payout 2026",
    "free btc claim bot 2026 no deposit",
    "bitcoin earning website no kyc withdraw 2026",
    # -- ETHEREUM --
    "free ethereum faucet 2026 instant withdrawal",
    "eth faucet auto claim free eth 2026",
    "ethereum mining free cloud 2026 no deposit",
    "free eth every hour claim 2026 no survey",
    "ethereum earning sites free withdraw 2026",
    "eth staking rewards free 2026 no minimum",
    # -- LITECOIN --
    "free ltc faucet instant withdrawal 2026 auto pay",
    "litecoin mining sites free 2026 no deposit cpu",
    "ltc auto claim bot free litecoin 2026",
    "free litecoin every hour claim faucet 2026",
    "litecoin faucet instant pay wallet 2026 no kyc",
    # -- DOGECOIN --
    "free dogecoin faucet instant withdrawal 2026",
    "doge faucet auto claim free doge 2026 no survey",
    "dogecoin mining free sites 2026 no deposit",
    "free doge every hour claim 2026 instant payout",
    "doge coin earning sites free withdraw wallet 2026",
    # -- SOLANA --
    "free solana faucet 2026 claim free sol",
    "solana earning sites free 2026 no deposit",
    "solana airdrop claim free tokens 2026",
    "solana staking rewards free 2026 no minimum",
    "free sol every hour claim faucet 2026",
    # -- USDT / USDC --
    "free usdt faucet 2026 claim free tether",
    "usdt earning sites free withdrawal 2026",
    "tether usdt staking rewards free 2026",
    "free usdc claim sites 2026 no deposit",
    "earn usdt by clicking 2026 no survey",
    # -- TRON --
    "free tron faucet 2026 claim free trx",
    "trx earning sites free 2026 instant withdrawal",
    "free trx every hour claim 2026 no deposit",
    "tron staking earn free trx 2026",
    # -- XRP --
    "free xrp faucet 2026 claim free ripple",
    "xrp earning sites free 2026 no deposit",
    "free xrp every hour claim 2026 instant payout",
    # -- BNB --
    "free bnb faucet 2026 claim free binance coin",
    "bnb earning sites free 2026 no deposit",
    "free bnb every hour smart chain 2026",
    # -- MATIC / POLYGON --
    "free matic faucet 2026 polygon claim free",
    "polygon staking earn free matic 2026",
    "free polygon matic every hour 2026 no deposit",
    # -- ADA --
    "free cardano faucet 2026 claim free ada",
    "ada staking earn free cardano 2026",
    "free ada every hour claim 2026",
    # -- USD / FIAT --
    "earn usd free 2026 no deposit instant payout paypal",
    "get paid in usd for simple tasks 2026 no survey",
    "earn dollars online free 2026 automated",
    "usd earning sites free registration 2026 no fees",
    "earn money paypal instant withdrawal 2026 free",
    "micro tasks pay usd 2026 no kyc no id",
    "click earn usd cash 2026 free automated",
    "earn paypal free 2026 every day automated",
    "passive income usd 2026 no investment free",
    # -- FAUCET AGGREGATORS --
    "best crypto faucets 2026 free btc eth ltc doge sol",
    "multi crypto faucet claim multiple coins 2026",
    "crypto faucet list 2026 instant withdrawal all coins",
    "auto claim crypto faucet bot multiple coins 2026",
    "faucet pay crypto instant free claim all coins",
    # -- CORE LTC/MINING --
    "crypto mining faucet earn free btc eth ltc doge 2026",
    "auto claim crypto faucet bot 2026",
    # -- HYPER-TARGETED --
    "one click crypto faucet instant pay 2026",
    "crypto faucet instant pay to wallet 2026",
    "crypto claim bot no survey no tasks 2026",
    "micro crypto faucet instant payout 2026",
    # -- CLICK TO EARN / PTC --
    "paid to click sites pay instantly crypto 2026 no survey",
    "click one button earn bitcoin free 2026",
    "ptc sites instant withdrawal crypto 2026",
    "best paid to click sites 2026 crypto payout",
    "click and earn crypto no tasks free 2026",
    # -- TELEGRAM BOTS --
    "telegram earning bot crypto free 2026 withdrawal",
    "telegram faucet bot crypto 2026 btc eth sol",
    "telegram mining bot free payout 2026",
    "telegram crypto claim bot instant 2026",
    "best telegram earning bots 2026 no investment",
    "telegram auto earn bot free withdrawal 2026",
    # -- AIRDROPS --
    "new crypto airdrops 2026 free tokens claim",
    "solana airdrop 2026 claim free",
    "telegram bot airdrop claim 2026",
    "bitcoin airdrop free 2026 claim btc",
    "ethereum airdrop 2026 free eth claim",
    # -- PASSIVE / DEFI --
    "passive income crypto staking defi 2026 no minimum",
    "free crypto staking rewards 2026 multi chain",
    "defi yield farming 2026 no minimum deposit",
    "passive crypto income 2026 set and forget",
    # -- AUTOMATION TOOLS --
    "browser automation earn crypto free 2026",
    "telegram bot earn crypto 2026 free automated",
    "auto trading crypto bot free 2026",
    "free crypto arbitrage bot 2026",
    "auto claim bot faucet multi coin 2026",
    # -- CASHBACK / AFFILIATE --
    "best cashback apps 2026 free money crypto",
    "affiliate programs crypto 2026 high paying free",
    # -- GENERAL --
    "earn free crypto no deposit 2026 withdraw instantly",
    "automatic crypto earning platform 2026 no investment",
    "free bitcoin earning sites 2026 withdraw to wallet",
    "cloud mining free trial 2026 no deposit btc",
    "web3 earn crypto free 2026 browser mining",
    "telegram mining bot free withdrawal 2026",
    "faucet pay crypto instant 2026 free claim",
    # -- NO-CAPTCHA QUERIES (appended) --
    'crypto faucet no captcha instant withdrawal 2026',
    'no captcha bitcoin faucet instant pay 2026',
    'free crypto no captcha no survey no tasks 2026',
    'no captcha faucet faucetpay 2026',
    'auto claim crypto no captcha 2026',
    'earn crypto without captcha 2026 free',
    'no captcha trx faucet instant payout 2026',
    'no captcha ltc faucet free litecoin 2026',
    'no captcha doge faucet 2026 instant',
    'no captcha usdt faucet 2026 free tether',
    'auto faucet crypto no manual claim 2026',
    'auto claim bitcoin bot no captcha 2026',
    'telegram auto earn crypto bot no captcha 2026',
    'auto faucet site no verification 2026 faucetpay',
    'set and forget crypto faucet 2026',
    'auto claim crypto platform no captcha 2026',
    'automatic crypto faucet faucetpay 2026',
    'one click crypto faucet no captcha 2026',
    'simple captcha crypto faucet faucetpay 2026',
    'text captcha faucet free crypto 2026',
    'easy captcha bitcoin faucet instant withdrawal 2026',
    'math captcha crypto faucet free 2026',
    'basic captcha faucet instant faucetpay 2026',
    'anti bot simple captcha crypto faucet 2026',
    'ptc sites instant withdrawal no captcha 2026 faucetpay',
    'click earn crypto no captcha instant payout 2026',
    'paid to click crypto faucetpay no captcha 2026',
    'shortlink earn crypto no captcha faucetpay 2026',
    'multi coin auto faucet faucetpay 2026',
    'all in one crypto faucet auto claim 2026 faucetpay',
    'bitcoin litecoin doge auto faucet faucetpay 2026 no captcha',
    'telegram crypto faucet bot auto earn faucetpay 2026',
    'telegram earning bot no captcha free withdrawal 2026',
    'best telegram faucet bot faucetpay 2026 no captcha',
    'telegram mining bot auto claim faucetpay 2026',
    'free crypto earning platform no captcha 2026 withdraw',
    'earn free crypto automatic faucetpay 2026',
    'crypto faucet list no captcha 2026',
    'claim free crypto auto faucet instant faucetpay',
    'best paying crypto faucet no captcha 2026',
]

# === HELPER FUNCTIONS ===

def extract_domain(url: str) -> str:
    try:
        parsed = urllib.parse.urlparse(url)
        domain = parsed.netloc or parsed.hostname or ""
        domain = domain.lower()
        if domain.startswith("www."):
            domain = domain[4:]
        return domain
    except:
        return url.lower() if url else ""

SUPICIOUS_URL_PATTERNS = [
    "bit.ly", "tinyurl", "shorturl", "short.link", "shorte.st",
    "adf.ly", "ouo.io", "shortener", "tiny.cc", "t.co",
    "goo.gl", "is.gd", "buff.ly", "ow.ly", "rebrand.ly",
    "bc.vc", "linkbucks", "ceesty", "clkme", "shrinkearn", "shortest",
    "click", "track", "redirect?", "url?q=", "out?",
]

def is_safe_url(url: str) -> bool:
    if not url:
        return False
    url_lower = url.lower()
    for pattern in SUPICIOUS_URL_PATTERNS:
        if pattern in url_lower:
            return False
    if not url_lower.startswith("http"):
        return False
    dangerous_exts = [".exe", ".dll", ".bat", ".scr", ".msi", ".zip", ".rar", ".apk", ".vbs"]
    for ext in dangerous_exts:
        if url_lower.endswith(ext):
            return False
    return True

def resolve_url(url: str) -> str:
    if 'bing.com/ck/a' not in url:
        return url
    try:
        parsed = urllib.parse.urlparse(url)
        qs = urllib.parse.parse_qs(parsed.query)
        u_param = qs.get('u', [''])[0]
        if not u_param:
            return url
        for prefix_len in range(0, 5):
            encoded = u_param[prefix_len:]
            if not encoded:
                continue
            try:
                padding = 4 - len(encoded) % 4
                if padding != 4:
                    encoded_padded = encoded + '=' * padding
                else:
                    encoded_padded = encoded
                decoded = base64.b64decode(encoded_padded).decode('utf-8')
                if decoded.startswith('http'):
                    return decoded
            except:
                continue
    except:
        pass
    return url

def extract_ddg_url(u: str) -> str:
    if "duckduckgo.com/l/" in u:
        parsed = urllib.parse.urlparse(u)
        qs = urllib.parse.parse_qs(parsed.query)
        return qs.get("uddg", [u])[0]
    return u

# === AI GENERATION FUNCTIONS ===

def groq_generate(prompt_str, model=None):
    global _groq_key_index
    if not GROQ_API_KEYS:
        print("[Groq] No API keys configured")
        return None
    if model is None:
        model = DEFAULT_GROQ_MODEL
    for _ in range(len(GROQ_API_KEYS)):
        key = GROQ_API_KEYS[_groq_key_index]
        try:
            headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
            body = {
                "model": model,
                "messages": [{"role": "user", "content": prompt_str}],
                "temperature": 0.2,
                "max_tokens": 600
            }
            r = requests.post("https://api.groq.com/openai/v1/chat/completions",
                              headers=headers, json=body, timeout=30)
            data = r.json()
            if r.status_code == 200:
                return data["choices"][0]["message"]["content"]
            err_msg = data.get("error", {}).get("message", "")
            if "429" in str(r.status_code) or "rate_limit" in err_msg.lower():
                print(f"[Groq] Key {_groq_key_index+1}/{len(GROQ_API_KEYS)} rate limited, trying next...")
                _groq_key_index = (_groq_key_index + 1) % len(GROQ_API_KEYS)
                time.sleep(1)
                continue
            print(f"[Groq] Key {_groq_key_index+1}/{len(GROQ_API_KEYS)} error {r.status_code}: {err_msg[:80]}")
            _groq_key_index = (_groq_key_index + 1) % len(GROQ_API_KEYS)
        except requests.exceptions.Timeout:
            print(f"[Groq] Key {_groq_key_index+1}/{len(GROQ_API_KEYS)} timeout, trying next...")
            _groq_key_index = (_groq_key_index + 1) % len(GROQ_API_KEYS)
            time.sleep(1)
        except Exception as e:
            print(f"[Groq] Key {_groq_key_index+1}/{len(GROQ_API_KEYS)} exception: {str(e)[:60]}")
            _groq_key_index = (_groq_key_index + 1) % len(GROQ_API_KEYS)
            time.sleep(1)
    print("[Groq] All API keys exhausted")
    return None

def openrouter_generate(prompt_str, model=None):
    global _openrouter_key_index
    if not OPENROUTER_API_KEYS:
        print("[OpenRouter] No API keys configured")
        return None
    if model is None:
        model = "openai/gpt-4o-mini"
    for _ in range(len(OPENROUTER_API_KEYS)):
        key = OPENROUTER_API_KEYS[_openrouter_key_index]
        try:
            headers = {
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/json",
                "HTTP-Referer": "https://github.com/wo312092-creator/opportunity-hunter-v2",
                "X-Title": "Opportunity Hunter V2"
            }
            body = {
                "model": model,
                "messages": [{"role": "user", "content": prompt_str}],
                "temperature": 0.2,
                "max_tokens": 800
            }
            r = requests.post("https://openrouter.ai/api/v1/chat/completions",
                              headers=headers, json=body, timeout=30)
            data = r.json()
            if r.status_code == 200:
                return data["choices"][0]["message"]["content"]
            err_msg = data.get("error", {}).get("message", "")
            if "429" in str(r.status_code) or "rate" in err_msg.lower():
                print(f"[OpenRouter] Key {_openrouter_key_index+1}/{len(OPENROUTER_API_KEYS)} rate limited, trying next...")
                _openrouter_key_index = (_openrouter_key_index + 1) % len(OPENROUTER_API_KEYS)
                time.sleep(1)
                continue
            print(f"[OpenRouter] Key {_openrouter_key_index+1} error {r.status_code}: {err_msg[:60]}")
            _openrouter_key_index = (_openrouter_key_index + 1) % len(OPENROUTER_API_KEYS)
        except requests.exceptions.Timeout:
            print(f"[OpenRouter] Key {_openrouter_key_index+1} timeout, trying next...")
            _openrouter_key_index = (_openrouter_key_index + 1) % len(OPENROUTER_API_KEYS)
            time.sleep(1)
        except Exception as e:
            print(f"[OpenRouter] Key {_openrouter_key_index+1} exception: {str(e)[:60]}")
            _openrouter_key_index = (_openrouter_key_index + 1) % len(OPENROUTER_API_KEYS)
            time.sleep(1)
    print("[OpenRouter] All API keys exhausted")
    return None

def gemini_generate(model, prompt_str):
    global _gemini_key_index
    if not GEMINI_API_KEYS:
        raise Exception("No Gemini API keys configured")
    for _ in range(len(GEMINI_API_KEYS)):
        key = GEMINI_API_KEYS[_gemini_key_index]
        try:
            genai.configure(api_key=key)
            resp = model.generate_content(prompt_str)
            return resp
        except Exception as e:
            if "429" not in str(e):
                raise
            print(f"[Gemini] Key {_gemini_key_index+1}/{len(GEMINI_API_KEYS)} quota exceeded, trying next...")
            _gemini_key_index = (_gemini_key_index + 1) % len(GEMINI_API_KEYS)
            time.sleep(1)
    raise Exception("All Gemini API keys exhausted")

# === SCORING ===

PLUMBING_FAUCET_KEYWORDS = ["kitchen","bathroom","sink","plumbing","delta","lowes","homedepot","home depot","ferguson","moen","toilet","shower","plumber","vanity","pipe","tub","spout","cartridge","hardware store","plumbing supply"]

CLASSIFY_KEYWORDS = [
    (["faucet","free crypto","claim","btc faucet","eth faucet","crypto faucet","bitcoin faucet"], "Crypto Faucet", ["faucet","crypto","free"]),
    (["airdrop","token distribution","free token","claim airdrop","crypto airdrop"], "Airdrop", ["airdrop","crypto","free"]),
    (["ptc","paid-to-click","bux","click ads","get paid to","paidtoclick","earn per click"], "PTC / GPT", ["ptc","gpt","click"]),
    (["survey","paid survey","market research","paid surveys"], "Paid Surveys", ["survey","research"]),
    (["cashback","cash back","rebate","shopping","cashback site"], "Cashback", ["cashback","shopping"]),
    (["affiliate","referral","refer","affiliate program"], "Affiliate", ["affiliate","referral"]),
    (["stake","staking","defi","yield","lend","apr","liquidity","pool"], "DeFi / Staking", ["defi","staking","yield"]),
    (["play to earn","p2e","gamefi","nft game","play-to-earn","crypto game"], "Play-to-Earn", ["p2e","gaming","nft"]),
    (["mining","cloud mining","hash","mine","miner","crypto mining","ltc miner","bitcoin miner","free mining","mining pool"], "Mining", ["mining","crypto"]),
    (["trading bot","auto trade","signal","copy trade","trading platform","grid trading"], "Trading", ["trading","bot","automation"]),
    (["micro task","microtask","data entry","freelance","micro job","gig"], "Micro Tasks", ["micro-task","freelance"]),
    (["browser automation","auto earn","auto claim","auto bot","automation bot","auto click","auto faucet"], "Automation Bot", ["automation","bot"]),
    (["earn crypto","free crypto","get crypto","crypto earn","crypto reward","crypto bonus"], "Crypto Earnings", ["crypto","earn"]),
    (["ltc","litecoin","dogecoin","doge coin","doge mining","ltc mining","free ltc","free doge"], "Altcoin Mining", ["mining","ltc","doge"]),
]

def rule_score_automation(title: str, desc: str, url: str) -> tuple:
    c = f"{title} {desc} {url}".lower()
    if any(s in c for s in ["dictionary","meaning","definition","wikipedia","encyclopedia","tutorial","course","educational","academic"]):
        return 1, "Educational content", "", ""
    bonus, signals = 0, 0
    for words, pts in [
        (["mining","miner","hash","cloud mining","free btc","free eth","free ltc","free doge"], 3),
        (["faucet","claim","withdraw","crypto faucet"], 3),
        (["auto","bot","telegram bot","auto claim","auto bot"], 2),
        (["click","ptc","paid-to-click","earn per"], 2),
        (["passive income","staking","stake","passive"], 2),
        (["refer","affiliate","earn crypto","reward","bonus","cash"], 1)]:
        if any(w in c for w in words):
            if words[0] == "faucet" and any(pk in c for pk in PLUMBING_FAUCET_KEYWORDS):
                continue
            bonus += pts
            signals += 1
    if any(w in c for w in ["survey", "data entry", "freelance", "offer wall", "offerwall", "gpt site", "micro task", "app download", "complete offers"]):
        bonus = -10
    score = min(10, max(0, 5 + bonus // max(1, signals // 2 if signals else 1)))
    if signals == 0 or bonus < -5:
        if bonus < -5:
            return 0, "Anti-pattern: survey/GPT site requires human work, not automatable", "", ""
        return 0, "Low: no automation signals", "", ""
    how_to_earn = "Visit website, complete tasks/claims, withdraw earnings to wallet"
    how_to_automate = "Use Playwright browser automation on GitHub Actions to login, claim, and withdraw on schedule"
    if score >= 7:
        return score, "High automation: bot-friendly signals", how_to_earn, how_to_automate
    elif score >= 4:
        return score, "Medium: partially automatable", how_to_earn, how_to_automate
    return score, "Low: human interaction needed", how_to_earn, how_to_automate

def score_automation_llm(title: str, desc: str, url: str) -> tuple:
    prompt = f"""Rate this opportunity's AUTOMATION POTENTIAL (0-10).
High (7-10): fully automatable with a GitHub bot - auto-click, auto-claim, auto-mining, auto-faucet, auto-task scripts. No human needed.
Medium (4-6): partially automatable - needs occasional captchas or approvals.
Low (0-3): requires continuous human work - typing, reading, manual trading.

Title: {title[:200]}
Description: {desc[:300]}
URL: {url[:200]}

Respond ONLY with JSON: {{"score": N, "reason": "short reason", "how_to_earn": "how to earn from this", "how_to_automate": "how to automate with GitHub Actions"}}"""

    text = groq_generate(prompt, model=FAST_GROQ_MODEL)
    if text:
        try:
            match = re.search(r'\{[^}]+\}', text)
            if match:
                data = json.loads(match.group())
                score = max(0, min(10, int(data.get("score", 5))))
                return score, str(data.get("reason", ""))[:200], str(data.get("how_to_earn", ""))[:300], str(data.get("how_to_automate", ""))[:300]
        except Exception:
            pass
    if not text and OPENROUTER_API_KEYS:
        print(f"[OpenRouter] Trying scoring...", end=" ", flush=True)
        text = openrouter_generate(prompt, model="openai/gpt-4o-mini")
        if text:
            try:
                match = re.search(r'\{[^}]+\}', text)
                if match:
                    data = json.loads(match.group())
                    score = max(0, min(10, int(data.get("score", 5))))
                    return score, str(data.get("reason", ""))[:200], str(data.get("how_to_earn", ""))[:300], str(data.get("how_to_automate", ""))[:300]
            except Exception:
                pass
    if not text and GEMINI_API_KEYS:
        try:
            model = genai.GenerativeModel('gemini-2.0-flash')
            resp = gemini_generate(model, prompt)
            text = resp.text.strip()
            match = re.search(r'\{[^}]+\}', text)
            if match:
                data = json.loads(match.group())
                score = max(0, min(10, int(data.get("score", 5))))
                return score, str(data.get("reason", ""))[:200], str(data.get("how_to_earn", ""))[:300], str(data.get("how_to_automate", ""))[:300]
        except Exception as e:
            if "quota" in str(e).lower() or "429" in str(e):
                print(f"[Gemini] All API keys exhausted")
            else:
                print(f"[Gemini] Error: {e}")
    return rule_score_automation(title, desc, url)

def classify(item: dict, genai_scoring: bool = True) -> Optional['Opportunity']:
    t, u, d = item.get("title",""), item.get("url",""), item.get("description","")
    if not t or not u:
        return None
    c = f"{t} {d}".lower()
    for keywords, cat, tags in CLASSIFY_KEYWORDS:
        if any(k in c for k in keywords):
            if cat == "Crypto Faucet" and any(pk in c for pk in PLUMBING_FAUCET_KEYWORDS):
                continue
            if genai_scoring:
                auto_score, auto_reason, how_to_earn, how_to_automate = score_automation_llm(t, d, u)
            else:
                auto_score, auto_reason, how_to_earn, how_to_automate = rule_score_automation(t, d, u)
            fea = "Easy" if auto_score >= 7 else "Medium" if auto_score >= 4 else "Hard"
            return Opportunity(id=hashlib.md5(f"{t}{u}{time.time()}".encode()).hexdigest()[:12],
                title=t[:120], url=u[:300], category=cat, description=d[:500],
                profit_per_hour="Analyzing...", profit_per_day="Analyzing...", profit_per_week="Analyzing...",
                profit_per_month="Analyzing...", profit_per_year="Analyzing...",
                effort_level="Analyzing...", automation_potential=auto_score, automation_reason=auto_reason,
                how_to_earn=how_to_earn, how_to_automate=how_to_automate, feasibility=fea,
                source="web_search", found_date=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                tags=tags, status="new")
    if any(k in c for k in ["earn","money","income","profit","passive","free","reward","bonus","cash","crypto","bitcoin"]):
        if genai_scoring:
            auto_score, auto_reason, how_to_earn, how_to_automate = score_automation_llm(t, d, u)
        else:
            auto_score, auto_reason, how_to_earn, how_to_automate = rule_score_automation(t, d, u)
        fea = "Easy" if auto_score >= 7 else "Medium" if auto_score >= 4 else "Hard"
        return Opportunity(id=hashlib.md5(f"{t}{u}{time.time()}".encode()).hexdigest()[:12],
            title=t[:120], url=u[:300], category="General", description=d[:500],
            profit_per_hour="Analyzing...", profit_per_day="Analyzing...", profit_per_week="Analyzing...",
            profit_per_month="Analyzing...", profit_per_year="Analyzing...",
            effort_level="Analyzing...", automation_potential=auto_score, automation_reason=auto_reason,
            how_to_earn=how_to_earn, how_to_automate=how_to_automate, feasibility=fea,
            source="web_search", found_date=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            tags=["money","earn"], status="new")
    return None
# === SEARCH FUNCTIONS ===

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
]

def search_bing_html(query: str, ua_idx: int = 0) -> list:
    try:
        resp = requests.get("https://www.bing.com/search", params={"q": query, "count": 10},
            headers={"User-Agent": USER_AGENTS[ua_idx % len(USER_AGENTS)], "Accept-Language": "en-US,en;q=0.9"},
            timeout=10)
        links = re.findall(r'<a[^>]+href="(https?://[^"]+)"[^>]*><h2>(.*?)</h2>', resp.text, re.DOTALL)
        if not links:
            links = re.findall(r'<h2[^>]*><a[^>]+href="(https?://[^"]+)"[^>]*>(.*?)</a></h2>', resp.text, re.DOTALL)
        snippets = re.findall(r'<p[^>]*>(.*?)</p>', resp.text, re.DOTALL)
        results = []
        for i, (u, t) in enumerate(links[:10]):
            desc = html_mod.unescape(re.sub(r'<[^>]+>', '', snippets[i] if i < len(snippets) else "")).strip()[:300] if i < len(snippets) else ""
            results.append({"title": html_mod.unescape(re.sub(r'<[^>]+>', '', t)).strip()[:200], "url": u, "description": desc})
        return results
    except Exception as e:
        print(f"[Bing HTML] Error: {e}")
        return []

def search_ddg(query: str, ua_idx: int = 0) -> list:
    try:
        resp = requests.get("https://html.duckduckgo.com/html/", params={"q": query},
            headers={"User-Agent": USER_AGENTS[ua_idx % len(USER_AGENTS)]}, timeout=8)
        links = re.findall(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', resp.text, re.DOTALL)
        snippets = re.findall(r'<a[^>]+class="result__snippet"[^>]*>(.*?)</a>', resp.text, re.DOTALL)
        if not links:
            return []
        results = []
        for i, (u, t) in enumerate(links[:10]):
            desc = html_mod.unescape(re.sub(r'<[^>]+>', '', snippets[i] if i < len(snippets) else "")).strip()[:300] if i < len(snippets) else ""
            url = extract_ddg_url(u)
            if url:
                results.append({"title": html_mod.unescape(re.sub(r'<[^>]+>', '', t)).strip()[:200], "url": url, "description": desc})
        return results
    except:
        return []

def search_startpage(query: str, ua_idx: int = 0) -> list:
    try:
        resp = requests.post("https://www.startpage.com/sp/search",
            data={"query": query, "language": "en", "cat": "web", "page": 1},
            headers={"User-Agent": USER_AGENTS[ua_idx % len(USER_AGENTS)]}, timeout=10)
        if resp.status_code == 200:
            links = re.findall(r'<a[^>]+class="w-gl__result-title"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', resp.text, re.DOTALL)
            descs = re.findall(r'<p class="w-gl__description[^>]*>(.*?)</p>', resp.text, re.DOTALL)
            results = []
            for i, (u, t) in enumerate(links[:10]):
                desc = html_mod.unescape(re.sub(r'<[^>]+>', '', descs[i] if i < len(descs) else "")).strip()[:300]
                results.append({"title": html_mod.unescape(re.sub(r'<[^>]+>', '', t)).strip()[:200], "url": u, "description": desc})
            return results
        return []
    except:
        return []

def search_yahoo(query: str, ua_idx: int = 0) -> list:
    try:
        resp = requests.get("https://search.yahoo.com/search", params={"p": query, "n": 10},
            headers={"User-Agent": USER_AGENTS[ua_idx % len(USER_AGENTS)]}, timeout=10)
        if resp.status_code != 200:
            return []
        links = re.findall(r'<a[^>]+class="[^"]*ac-algo[^"]*"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', resp.text, re.DOTALL)
        if not links:
            links = re.findall(r'<h3[^>]*><a[^>]+href="([^"]+)"[^>]*>(.*?)</a></h3>', resp.text, re.DOTALL)
        snippets = re.findall(r'<div[^>]*class="[^"]*compText[^"]*"[^>]*>(.*?)</div>', resp.text, re.DOTALL)
        if not snippets:
            snippets = re.findall(r'<p[^>]*class="[^"]*[Ff]c-[^"]*"[^>]*>(.*?)</p>', resp.text, re.DOTALL)
        results = []
        for i, (u, t) in enumerate(links[:10]):
            desc = html_mod.unescape(re.sub(r'<[^>]+>', '', snippets[i] if i < len(snippets) else "")).strip()[:300]
            results.append({"title": html_mod.unescape(re.sub(r'<[^>]+>', '', t)).strip()[:200], "url": u, "description": desc})
        return results
    except:
        return []

def search_reddit(query: str, ua_idx: int = 0) -> list:
    subreddits = ["beermoney", "cryptocurrency", "airdrops", "sidehustle", "cryptomining", "earncrypto", "btc"]
    keywords = ["ltc", "litecoin", "mining", "doge", "dogecoin", "btc", "bitcoin", "eth", "ethereum",
                "faucet", "airdrop", "claim", "earning", "earn", "free", "crypto", "solana", "auto"]
    chosen_q = query
    for kw in keywords:
        if kw in query.lower():
            chosen_q = query
            break
    else:
        chosen_q = "free crypto earn " + query.split()[-1] if len(query.split()) > 1 else query
    results = []
    seen_urls = set()
    try:
        for sub in subreddits[:4]:
            try:
                url = f"https://www.reddit.com/r/{sub}/search.json?q={urllib.parse.quote(chosen_q)}&restrict_sr=1&sort=new&limit=5&t=year"
                resp = requests.get(url, headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) opportunity-hunter-v2/1.0"
                }, timeout=10)
                if resp.status_code != 200:
                    continue
                data = resp.json()
                children = data.get("data", {}).get("children", [])
                for child in children[:5]:
                    try:
                        d = child.get("data", {})
                        url = d.get("url", "")
                        title = d.get("title", "")
                        permalink = d.get("permalink", "")
                        selftext = d.get("selftext", "")[:300]
                        if not url or not title:
                            continue
                        if url.startswith("/r/"):
                            url = "https://www.reddit.com" + url
                        url_key = url.split("?")[0].rstrip("/")
                        if url_key in seen_urls:
                            continue
                        seen_urls.add(url_key)
                        desc = selftext if selftext else f"Reddit r/{sub} - {permalink[:100]}"
                        results.append({"title": title.strip()[:200], "url": url, "description": desc[:300]})
                    except:
                        continue
            except:
                continue
    except:
        pass
    return results

def search_github(query: str, ua_idx: int = 0) -> list:
    github_queries = [
        query, "crypto earning automation", "mining bot script",
        "auto faucet claim", "LTC mining tool", "crypto auto withdraw",
    ]
    chosen_q = query
    kw = query.lower().split()
    if any(w in kw for w in ["mining", "miner"]):
        chosen_q = "crypto mining automation script"
    elif any(w in kw for w in ["faucet", "claim"]):
        chosen_q = "auto faucet claim bot"
    elif any(w in kw for w in ["airdrop", "air drop"]):
        chosen_q = "crypto airdrop bot"
    elif any(w in kw for w in ["bot", "telegram"]):
        chosen_q = "telegram crypto earn bot"
    elif any(w in kw for w in ["earning", "earn", "free"]):
        chosen_q = "crypto earning script automated"
    results = []
    seen_urls = set()
    try:
        repo_url = f"https://api.github.com/search/repositories?q={urllib.parse.quote(chosen_q)}+crypto+script&sort=updated&per_page=5"
        resp = requests.get(repo_url, headers={
            "Accept": "application/vnd.github.v3+json",
            "User-Agent": "opportunity-hunter-v2"
        }, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            for item in data.get("items", [])[:5]:
                try:
                    url = item.get("html_url", "")
                    title = item.get("full_name", item.get("name", ""))
                    desc = item.get("description", "") or ""
                    stars = item.get("stargazers_count", 0)
                    if not url or not title:
                        continue
                    url_key = url.split("?")[0].rstrip("/")
                    if url_key in seen_urls:
                        continue
                    seen_urls.add(url_key)
                    enriched_desc = f"{desc[:200]} {stars} stars" if desc else f"GitHub repo - {title} {stars} stars"
                    results.append({"title": title.strip()[:200], "url": url, "description": enriched_desc[:300]})
                except:
                    continue
        topic_url = f"https://api.github.com/search/repositories?q={urllib.parse.quote(chosen_q)}+topic:crypto&sort=updated&per_page=3"
        resp2 = requests.get(topic_url, headers={
            "Accept": "application/vnd.github.v3+json",
            "User-Agent": "opportunity-hunter-v2"
        }, timeout=10)
        if resp2.status_code == 200:
            data2 = resp2.json()
            for item in data2.get("items", [])[:3]:
                try:
                    url = item.get("html_url", "")
                    title = item.get("full_name", item.get("name", ""))
                    desc = item.get("description", "") or ""
                    if not url:
                        continue
                    url_key = url.split("?")[0].rstrip("/")
                    if url_key in seen_urls:
                        continue
                    seen_urls.add(url_key)
                    results.append({"title": title.strip()[:200], "url": url, "description": (desc[:200] if desc else "Crypto-related GitHub repo")[:300]})
                except:
                    continue
    except:
        pass
    return results

class PlaywrightPool:
    def __init__(self):
        self.browser = None
        self._pw = None
    def start(self):
        from playwright.sync_api import sync_playwright
        self._pw = sync_playwright()
        p = self._pw.start()
        self.browser = p.chromium.launch(headless=True, args=["--no-sandbox","--disable-setuid-sandbox","--window-size=1920,1080"])
    def search(self, query: str, q_idx: int = 0) -> list:
        if not self.browser:
            self.start()
        try:
            ctx = self.browser.new_context(
                user_agent=USER_AGENTS[q_idx % len(USER_AGENTS)],
                viewport={"width": 1920, "height": 1080},
                locale="en-US",
            )
            page = ctx.new_page()
            page.goto(f"https://www.bing.com/search?q={urllib.parse.quote(query)}&count=10", timeout=30000, wait_until="domcontentloaded")
            # Wait for results OR captcha to appear (avoids 'Page is navigating' on content())
            try:
                page.wait_for_selector("li.b_algo, #b_results, #captcha", timeout=15000)
            except:
                page.wait_for_timeout(3000)
            # Safely check for captcha — wrap content() in retry for navigation-safe reads
            for _attempt in range(3):
                try:
                    content = page.content()
                    break
                except Exception:
                    page.wait_for_timeout(2000)
            else:
                content = ""
            if "captcha" in content.lower() or "captcha" in page.url.lower():
                ctx.close()
                return []
            page.evaluate("window.scrollBy(0, 400)")
            page.wait_for_timeout(500)
            items = page.query_selector_all("li.b_algo")
            results = []
            for el in items[:12]:
                try:
                    h2 = el.query_selector("h2")
                    a = h2.query_selector("a[href^='http']") if h2 else None
                    if not a: continue
                    url = a.get_attribute("href") or ""
                    title = h2.inner_text().strip()
                    desc_el = el.query_selector("div.b_caption p, div.b_snippet")
                    desc = desc_el.inner_text()[:400] if desc_el else ""
                    if url and title:
                        results.append({"title": title, "url": url, "description": desc.strip()})
                except: continue
            ctx.close()
            return results
        except Exception as e:
            print(f"[Bing PW] Error: {e}")
            return []
    def search_google(self, query: str, q_idx: int = 0) -> list:
        if not self.browser:
            self.start()
        try:
            ctx = self.browser.new_context(
                user_agent=USER_AGENTS[q_idx % len(USER_AGENTS)],
                viewport={"width": 1920, "height": 1080},
                locale="en-US",
            )
            page = ctx.new_page()
            page.add_init_script("""() => {
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                Object.defineProperty(navigator, 'plugins', { get: () => [1,2,3,4,5] });
            }""")
            page.goto(f"https://www.google.com/search?q={urllib.parse.quote(query)}&num=10", timeout=20000)
            page.wait_for_timeout(2000)
            if "captcha" in page.content().lower() or "unusual traffic" in page.content().lower():
                print(f"[Google PW] Blocked", end=" ", flush=True)
                ctx.close()
                return []
            items = page.query_selector_all("div.g")
            results = []
            for el in items[:10]:
                try:
                    a = el.query_selector("a[href^='http']")
                    h3 = el.query_selector("h3")
                    if not a or not h3: continue
                    url = a.get_attribute("href") or ""
                    title = h3.inner_text().strip()
                    desc_el = el.query_selector("div.VwiC3b, span.st")
                    desc = desc_el.inner_text()[:400] if desc_el else ""
                    if url and title:
                        results.append({"title": title, "url": url, "description": desc.strip()})
                except: continue
            ctx.close()
            return results
        except Exception as e:
            print(f"[Google PW] Error: {e}")
            return []
    def visit_page(self, url: str) -> dict:
        if not self.browser:
            self.start()
        try:
            ctx = self.browser.new_context(
                user_agent=random.choice(USER_AGENTS),
                viewport={"width": 1920, "height": 1080},
                locale="en-US",
            )
            page = ctx.new_page()
            resp = page.goto(url, timeout=45000, wait_until="domcontentloaded")
            page.wait_for_timeout(3000)
            status = resp.status if resp else 0
            title = page.title()
            body_text = page.evaluate("() => document.body?.innerText?.substring(0, 8000) || ''")
            links = page.evaluate("""() => {
                const anchors = Array.from(document.querySelectorAll('a[href]'));
                return anchors.slice(0, 50).map(a => ({text: a.innerText?.trim()?.substring(0, 60) || '', href: a.href})).filter(x => x.text && x.href);
            }""")
            buttons = page.evaluate("""() => {
                const btns = Array.from(document.querySelectorAll('button, input[type="submit"], a[class*="btn"]'));
                return btns.slice(0, 30).map(b => ({text: b.innerText?.trim()?.substring(0, 40) || b.value?.substring(0, 40) || ''}));
            }""")
            inputs = page.evaluate("""() => {
                const inp = Array.from(document.querySelectorAll('input:not([type="hidden"])'));
                return inp.slice(0, 20).map(i => ({name: i.name || '', type: i.type || '', placeholder: i.placeholder || ''}));
            }""")
            ctx.close()
            return {
                "success": True, "status": status, "title": title,
                "body_text": body_text[:8000], "links": links,
                "buttons": [b["text"] for b in buttons if b["text"]], "inputs": inputs,
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
    def close(self):
        if self.browser:
            self.browser.close()
        if self._pw:
            try: self._pw.__exit__(None, None, None)
            except: pass

def curl_fetch_page(url: str) -> dict:
    try:
        try:
            from curl_cffi import requests as curl_requests
            r = curl_requests.get(url, impersonate='chrome', timeout=45,
                headers={"Accept-Language": "en-US,en;q=0.9"})
        except ImportError:
            import requests
            r = requests.get(url, timeout=45,
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"})
        text = r.text if hasattr(r, 'text') else (r.content.decode('utf-8', errors='replace') if hasattr(r, 'content') else '')
        title_m = re.search(r'<title[^>]*>(.*?)</title>', text, re.I | re.S)
        title = title_m.group(1).strip() if title_m else ""
        body_m = re.search(r'<body[^>]*>(.*?)</body>', text, re.I | re.S)
        body_html = body_m.group(1) if body_m else text
        body_text = re.sub(r'<script[^>]*>.*?</script>', '', body_html, flags=re.I | re.S)
        body_text = re.sub(r'<style[^>]*>.*?</style>', '', body_text, flags=re.I | re.S)
        body_text = re.sub(r'<[^>]+>', ' ', body_text)
        body_text = re.sub(r'\s+', ' ', body_text).strip()[:8000]
        buttons = re.findall(r'<button[^>]*>([^<]*)</button>', text, re.I)
        input_btns = re.findall(r'<input[^>]*type=["\']submit["\'][^>]*value=["\']([^"\']*)["\']', text, re.I)
        all_buttons = [b.strip() for b in buttons + input_btns if b.strip()]
        inputs = re.findall(r'<input[^>]*type=["\']?([^"\'\s>]+)["\']?[^>]*name=["\']([^"\']*)["\']', text, re.I | re.S)
        inputs2 = re.findall(r'<input[^>]*name=["\']([^"\']*)["\'][^>]*type=["\']?([^"\'\s>]+)', text, re.I | re.S)
        all_inputs = list(set([(t, n) for n, t in inputs] + [(t, n) for t, n in inputs2]))
        links = re.findall(r'<a[^>]*href=["\'](https?://[^"\']+)["\'][^>]*>([^<]*)</a>', text, re.I | re.S)
        return {
            "success": True,
            "status": r.status_code if hasattr(r, 'status_code') else (r.status if hasattr(r, 'status') else 0),
            "title": title, "body_text": body_text[:8000],
            "links": [{"text": t.strip()[:60], "href": h} for h, t in links[:50] if t.strip()],
            "buttons": all_buttons[:30],
            "inputs": [{"name": n, "type": t, "placeholder": ""} for t, n in all_inputs[:20]],
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

def search_exa(query: str) -> list:
    if not EXA_API_KEY:
        return []
    try:
        resp = requests.post(
            "https://api.exa.ai/search",
            headers={"x-api-key": EXA_API_KEY, "Content-Type": "application/json"},
            json={"query": query, "numResults": 10, "type": "auto", "useAutoprompt": True},
            timeout=15
        )
        if resp.status_code == 200:
            data = resp.json()
            results = []
            for r in data.get("results", []):
                url = r.get("url", "")
                title = r.get("title", "")
                desc = r.get("text", "")[:400]
                if url:
                    results.append({"title": title, "url": url, "description": desc})
            return results
        elif resp.status_code == 429:
            print("[Exa] Rate limited", end=" ", flush=True)
            return []
        else:
            print(f"[Exa] HTTP {resp.status_code}", end=" ", flush=True)
            return []
    except Exception as e:
        print(f"[Exa] Error: {e}", end=" ", flush=True)
        return []

def search_firecrawl(query: str) -> list:
    if not FIRECRAWL_API_KEY:
        return []
    session = requests.Session()
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    retries = Retry(total=3, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retries, pool_connections=4, pool_maxsize=8)
    session.mount("https://", adapter)
    try:
        resp = session.post(
            "https://api.firecrawl.dev/v1/search",
            headers={"Authorization": f"Bearer {FIRECRAWL_API_KEY}", "Content-Type": "application/json"},
            json={"query": query, "limit": 10, "scrapeOptions": {"formats": ["markdown"]}},
            timeout=20
        )
        if resp.status_code == 200:
            data = resp.json()
            results = []
            for r in data.get("data", []):
                url = r.get("url", "")
                title = r.get("title", "") or r.get("metadata", {}).get("title", "")
                desc_raw = r.get("markdown", "") or r.get("description", "") or ""
                desc = desc_raw[:400]
                if url:
                    results.append({"title": title, "url": url, "description": desc})
            return results
        elif resp.status_code == 429:
            print("[Firecrawl] Rate limited", end=" ", flush=True)
            return []
        else:
            print(f"[Firecrawl] HTTP {resp.status_code}", end=" ", flush=True)
            return []
    except Exception as e:
        print(f"[Firecrawl] Error: {e}", end=" ", flush=True)
        return []
    finally:
        session.close()

def search_all(query: str, pw: PlaywrightPool, q_idx: int) -> list:
    seen = set()
    results = []
    pw_results = pw.search(query, q_idx)
    print(f"[Bing PW] {len(pw_results)}", end=" ", flush=True)
    for r in pw_results:
        if r["url"] and r["url"] not in seen:
            r["url"] = resolve_url(r["url"])
            if r["url"] and r["url"] not in seen:
                seen.add(r["url"])
                results.append(r)
    time.sleep(random.uniform(0.3, 0.8))
    github_results = search_github(query, q_idx)
    print(f"[GitHub] {len(github_results)}", end=" ", flush=True)
    for r in github_results:
        if r["url"] and r["url"] not in seen:
            r["url"] = resolve_url(r["url"])
            if r["url"] and r["url"] not in seen:
                seen.add(r["url"])
                results.append(r)
    time.sleep(random.uniform(0.3, 0.8))
    exa_results = search_exa(query)
    print(f"[Exa] {len(exa_results)}", end=" ", flush=True)
    for r in exa_results:
        if r["url"] and r["url"] not in seen:
            if r["url"] not in seen:
                seen.add(r["url"])
                results.append(r)
    time.sleep(random.uniform(0.3, 0.8))
    fc_results = search_firecrawl(query)
    print(f"[Firecrawl] {len(fc_results)}", end=" ", flush=True)
    for r in fc_results:
        if r["url"] and r["url"] not in seen:
            if r["url"] not in seen:
                seen.add(r["url"])
                results.append(r)
    print(f"=> {len(results)} unique", flush=True)
    return results

# === CAPTCHA DETECTION (from NO-CAPTCHA Hunter) ===

def detect_captcha_type(url: str) -> dict:
    result = {
        "url": url, "has_captcha": False, "captcha_types": [],
        "has_faucetpay": False, "has_register": False, "has_login": False,
        "has_withdraw": False, "has_claim_button": False, "has_address_input": False,
        "title": "", "body_snippet": "", "status": 0, "success": False,
        "error": "", "captcha_sitekey": "", "detection_method": "",
    }
    try:
        from curl_cffi import requests as curl_requests
        r = curl_requests.get(url, impersonate='chrome', timeout=15,
            headers={"Accept-Language": "en-US,en;q=0.9"})
        text = r.text if hasattr(r, 'text') else r.content.decode('utf-8', errors='replace')
        result["status"] = r.status_code if hasattr(r, 'status_code') else 0
        result["success"] = True
        result["detection_method"] = "curl_cffi"
    except Exception as e:
        result["error"] = str(e)[:100]
        try:
            r = requests.get(url, timeout=15, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/125.0.0.0",
                "Accept-Language": "en-US,en;q=0.9"
            })
            text = r.text
            result["status"] = r.status_code
            result["success"] = True
            result["detection_method"] = "requests"
        except Exception as e2:
            result["error"] = str(e2)[:100]
            return result
    if not result["success"] or not text:
        return result
    m = re.search(r'<title[^>]*>(.*?)</title>', text, re.I | re.S)
    result["title"] = m.group(1).strip()[:120] if m else ""
    if re.search(r'google\.com/recaptcha/api\.js|g-recaptcha|data-sitekey=["\']', text, re.I):
        result["has_captcha"] = True
        result["captcha_types"].append("Google reCAPTCHA")
        m = re.search(r'data-sitekey=["\']([^"\']+)["\']', text)
        if m:
            result["captcha_sitekey"] = m.group(1)
    if re.search(r'hcaptcha\.com|data-hcaptcha-widget-id|h-captcha', text, re.I):
        result["has_captcha"] = True
        result["captcha_types"].append("hCaptcha")
    if re.search(r'challenges\.cloudflare\.com|cf-turnstile|turnstile', text, re.I):
        result["has_captcha"] = True
        result["captcha_types"].append("Cloudflare Turnstile")
    if re.search(r'solvemedia\.com|api-secure\.solvemedia', text, re.I):
        result["has_captcha"] = True
        result["captcha_types"].append("SolveMedia")
    if re.search(r'captcha\.(png|jpg|gif|jpeg)|simple-captcha|captcha_img', text, re.I):
        if not result["has_captcha"]:
            result["has_captcha"] = True
            result["captcha_types"].append("Text/Image captcha")
    if re.search(r'math.*captcha|what is.*\+|simple.*math|anti.*bot.*math', text, re.I):
        if "Text/Image captcha" not in result["captcha_types"]:
            result["has_captcha"] = True
            if "Math captcha" not in result["captcha_types"]:
                result["captcha_types"].append("Math captcha")
    if not result["has_captcha"]:
        if re.search(r'captcha|antibot|anti\.bot|human\.verif|are\.you\.human', text, re.I):
            result["has_captcha"] = True
            result["captcha_types"].append("Unknown (generic anti-bot)")
    if re.search(r'faucetpay|faucet\.pay|FaucetPay', text, re.I):
        result["has_faucetpay"] = True
    body = re.sub(r'<[^>]+>', ' ', text)
    body = re.sub(r'\s+', ' ', body).lower()
    if re.search(r'register|sign.?up|create.?account', body):
        result["has_register"] = True
    if re.search(r'login|sign.?in|log.?in', body):
        result["has_login"] = True
    if re.search(r'withdraw|payout|send to wallet|withdrawal', body):
        result["has_withdraw"] = True
    if re.search(r'claim|start|earn now|get free|claim now', body):
        result["has_claim_button"] = True
    if re.search(r'address|wallet.*input|enter.*address', body):
        result["has_address_input"] = True
    result["body_snippet"] = body[:500]
    return result

def deep_analyze_captcha(url: str, site_info: dict) -> dict:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return site_info
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"],
            )
            context = browser.new_context(
                user_agent=random.choice(USER_AGENTS),
                viewport={"width": 1280, "height": 800}, locale="en-US",
            )
            page = context.new_page()
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            time.sleep(3)
            rc_count = page.locator('iframe[title="reCAPTCHA"]').count()
            turn_count = page.locator('iframe[title*="Turnstile"]').count()
            hcap_count = page.locator('iframe[title*="hcaptcha"]').count()
            if rc_count > 0:
                site_info["has_captcha"] = True
                if "Google reCAPTCHA" not in site_info["captcha_types"]:
                    site_info["captcha_types"].append("Google reCAPTCHA")
                site_info["recaptcha_iframes"] = rc_count
                try:
                    anchor = page.frame_locator('iframe[title="reCAPTCHA"]')
                    anchor.locator(".recaptcha-checkbox-border").click(timeout=5000)
                    time.sleep(2)
                    challenge = page.locator('iframe[title*="challenge"]')
                    if challenge.count() > 0:
                        cf = page.frame_locator('iframe[title*="challenge"]')
                        challenge_text = cf.locator("body").inner_text(timeout=3000)[:200]
                        site_info["recaptcha_challenge"] = challenge_text
                        audio_btn = cf.locator("#recaptcha-audio-button")
                        if audio_btn.count() > 0 and audio_btn.first.is_visible():
                            site_info["captcha_has_audio"] = True
                            audio_btn.click(timeout=3000)
                            time.sleep(2)
                            audio_src = cf.locator("#audio-source")
                            src = audio_src.get_attribute("src", timeout=5000)
                            site_info["captcha_audio_source"] = bool(src)
                except:
                    pass
            if turn_count > 0:
                site_info["has_captcha"] = True
                if "Cloudflare Turnstile" not in site_info["captcha_types"]:
                    site_info["captcha_types"].append("Cloudflare Turnstile")
            if hcap_count > 0:
                site_info["has_captcha"] = True
                if "hCaptcha" not in site_info["captcha_types"]:
                    site_info["captcha_types"].append("hCaptcha")
            try:
                claim_btn = page.query_selector('#claimBtn, button:has-text("Claim"), button:has-text("Start"), button:has-text("Earn")')
                if claim_btn:
                    site_info["has_claim_button"] = True
                    site_info["claim_button_text"] = claim_btn.inner_text()[:40]
                    site_info["claim_button_disabled"] = claim_btn.is_disabled()
            except:
                pass
            try:
                addr = page.query_selector('#address, input[placeholder*="address" i], input[type="email"], input[placeholder*="email" i]')
                if addr:
                    site_info["has_address_input"] = True
                    site_info["address_placeholder"] = addr.get_attribute("placeholder") or ""
            except:
                pass
            browser.close()
    except Exception as e:
        site_info["playwright_error"] = str(e)[:100]
    return site_info

def calculate_automation_score(site_info: dict, title: str, body: str) -> dict:
    score = 5
    factors = []
    positives = []
    negatives = []
    captcha_types = [c.lower() for c in site_info.get("captcha_types", [])]
    if not site_info.get("has_captcha"):
        score += 4
        factors.append("NO_CAPTCHA=+4")
        positives.append("No captcha detected!")
    elif any("turnstile" in c for c in captcha_types):
        score += 2
        factors.append("TURNSTILE=+2")
        positives.append("Cloudflare Turnstile (bypassable)")
    elif any("text" in c or "image" in c or "math" in c for c in captcha_types):
        score += 1
        factors.append("SIMPLE_CAPTCHA=+1")
        positives.append("Simple captcha (OCR-solvable)")
    elif any("recaptcha" in c for c in captcha_types):
        challenge = site_info.get("recaptcha_challenge", "").lower()
        if "skip" in challenge or "none" in challenge:
            score += 1
            factors.append("RECAPTCHA_SKIP=+1")
            positives.append("reCAPTCHA with SKIP option")
        else:
            score -= 3
            factors.append("RECAPTCHA_IMAGE=-3")
            negatives.append("Google reCAPTCHA image challenge (hard to automate)")
    elif any("unknown" in c for c in captcha_types):
        score += 0
        factors.append("UNKNOWN_CAPTCHA=+0")
    if site_info.get("has_faucetpay"):
        score += 2
        factors.append("FAUCETPAY=+2")
        positives.append("Pays via FaucetPay (instant!)")
    if "auto" in body.lower() and any(w in body.lower() for w in ["auto claim", "auto faucet", "automated", "auto-earn"]):
        score += 2
        factors.append("AUTO_CLAIM=+2")
        positives.append("Has auto-claim feature")
    body_lower = body.lower()
    if any(kw in body_lower for kw in ["cloudflare", "turnstile", "recaptcha"]):
        score -= 1
        factors.append("GHA_RISK=-1")
        negatives.append("May block datacenter IPs")
    else:
        score += 1
        factors.append("GHA_FRIENDLY=+1")
        positives.append("GitHub Actions friendly")
    if site_info.get("has_address_input") and site_info.get("has_claim_button"):
        score += 1
        factors.append("SIMPLE_FLOW=+1")
        positives.append("Simple address + claim flow")
    if site_info.get("has_withdraw"):
        score += 1
        factors.append("HAS_WITHDRAW=+1")
        positives.append("Has withdrawal support")
    if "telegram" in body_lower and any(w in body_lower for w in ["bot", "@"]):
        score += 1
        factors.append("TELEGRAM=+1")
        positives.append("Telegram bot (easy to automate)")
    score = max(0, min(10, score))
    return {
        "score": score, "factors": factors, "positives": positives, "negatives": negatives,
        "verdict": "READY TO AUTOMATE" if score >= 7 else "POSSIBLE" if score >= 4 else "NOT RECOMMENDED",
    }

# === SITE TYPE CLASSIFICATION ===

def classify_site_type(url: str, title: str, body_text: str, page_data: dict) -> str:
    url_lower = url.lower()
    title_lower = title.lower()
    body_lower = body_text[:5000].lower() if body_text else ""
    links = page_data.get("links", []) if page_data else []
    if any(p in url_lower for p in ["blog/", "/guide", "medium.com", "wordpress", "blogspot", "substack"]):
        return "article"
    if any(p in title_lower for p in ["best ", "top ", "list of", "review", "guide", "how to ", "what is "]):
        return "article"
    if links:
        tg_links = sum(1 for l in links if "t.me/" in l.get("href", "") or "telegram" in l.get("href", "").lower())
        if len(links) > 0 and tg_links / len(links) > 0.7:
            return "telegram"
    exchange_domains = ["binance", "coinbase", "kraken", "kucoin", "bybit", "okx", "huobi", "gate.io", "crypto.com"]
    if any(d in url_lower for d in exchange_domains):
        return "exchange"
    if any(w in body_lower for w in ["mining pool", "pool mining", "hashrate", "stratum", "pool url"]):
        if "faucet" not in body_lower:
            return "mining_pool"
    if any(w in body_lower for w in ["paid to click", "ptc", "click ads", "earn per click", "bux"]):
        return "ptc"
    if any(w in body_lower for w in ["claim", "faucet", "free btc", "free ltc", "free doge", "free eth",
                                       "claim now", "withdraw to", "faucetpay", "satoshi", "micro wallet"]):
        return "faucet"
    if links:
        domain = extract_domain(url)
        external = sum(1 for l in links if extract_domain(l.get("href", "")) != domain)
        if len(links) > 10 and external / max(len(links), 1) > 0.5:
            return "directory"
    return "unknown"

# === SCRAPY INTEGRATION ===

def search_scrapy(query: str, existing_results: list, pw: PlaywrightPool) -> list:
    urls_to_crawl = [r["url"] for r in existing_results[:8] if r.get("url")]
    if not urls_to_crawl:
        return []
    print(f"[Scrapy] Deep crawling {len(urls_to_crawl)} pages...", flush=True)
    new_results = []
    seen_hrefs = set()
    for url in urls_to_crawl:
        try:
            page_data = curl_fetch_page(url)
            if not page_data.get("success"):
                print(f"  [Scrapy] Failed: {url[:60]}")
                continue
            print(f"  [Scrapy] OK: {url[:60]} -> {len(page_data.get('links', []))} links", flush=True)
            for link in page_data.get("links", []):
                href = link.get("href", "")
                text = link.get("text", "")
                if not href or not href.startswith("http"):
                    continue
                href_clean = href.split("?")[0].rstrip("/")
                if href_clean in seen_hrefs:
                    continue
                seen_hrefs.add(href_clean)
                combined = (text + " " + href).lower()
                if any(w in combined for w in ["faucet", "earn", "free", "crypto", "btc", "ltc", "doge",
                                                 "eth", "claim", "mining", "airdrop", "withdraw", "reward"]):
                    new_results.append({
                        "title": (text or "Deep crawl find")[:200],
                        "url": href,
                        "description": f"Found via deep crawl of {url[:60]}"
                    })
        except Exception as e:
            print(f"  [Scrapy] Error: {url[:60]} -> {str(e)[:50]}")
    print(f"[Scrapy] Found {len(new_results)} new earning links from deep crawl", flush=True)
    return new_results[:30]

# === OPPORTUNITY CLASS + SHEET COLUMNS ===

SHEET_COLUMNS = [
    "ID", "Website Name", "URL", "Category", "What It Does",
    "How To Earn", "Per Hour", "Per Day", "Per Week", "Per Month", "Per Year",
    "Auto Score", "How To Automate (GitHub)", "Feasibility", "Source",
    "Found Date", "Status", "Verification",
    "Captcha Type", "Site Type", "No-Captcha Score"
]

@dataclass
class Opportunity:
    id: str = ""
    title: str = ""
    url: str = ""
    category: str = ""
    description: str = ""
    profit_per_hour: str = ""
    profit_per_day: str = ""
    profit_per_week: str = ""
    profit_per_month: str = ""
    profit_per_year: str = ""
    effort_level: str = ""
    automation_potential: int = 0
    automation_reason: str = ""
    source: str = ""
    found_date: str = ""
    tags: list = field(default_factory=list)
    status: str = "new"
    how_to_earn: str = ""
    how_to_automate: str = ""
    feasibility: str = ""
    workflow_steps: str = ""
    tools_needed: str = ""
    automation_plan: str = ""
    site_analyzed: bool = False
    deep_analysis_score: int = 0
    captcha_type: str = ""
    site_type: str = ""
    no_captcha_score: int = 0

# === DEEP ANALYSIS ===

def deep_analyze_site(pw: PlaywrightPool, opp: Opportunity) -> Opportunity:
    print(f"\n[Deep] Analyzing: {opp.title[:50]}...", flush=True)
    if not is_safe_url(opp.url):
        print(f"[Deep] Skipped (unsafe/redirect URL)")
        opp.deep_analysis_score = 0
        opp.site_analyzed = True
        opp.effort_level = "Not analyzed - URL skipped for safety"
        opp.profit_per_hour = "N/A"; opp.profit_per_day = "N/A"
        opp.profit_per_week = "N/A"; opp.profit_per_month = "N/A"
        opp.profit_per_year = "N/A"
        return opp
    page_data = curl_fetch_page(opp.url)
    if page_data.get("success"):
        print(f"[Deep] curl_cffi OK: {page_data['status']}, {len(page_data.get('body_text',''))} chars")
    else:
        print(f"[Deep] curl_cffi failed ({page_data.get('error','unknown')}), trying Playwright...")
        page_data = pw.visit_page(opp.url)
    if not page_data.get("success"):
        print(f"[Deep] Cannot reach site: {page_data.get('error', 'unknown')}")
        opp.deep_analysis_score = 0; opp.site_analyzed = True
        opp.profit_per_hour = "N/A (site unreachable)"; opp.profit_per_day = "N/A"
        opp.profit_per_week = "N/A"; opp.profit_per_month = "N/A"; opp.profit_per_year = "N/A"
        return opp
    page_text = page_data.get("body_text", "")[:6000]
    buttons_text = ", ".join(page_data.get("buttons", [])[:15])
    inputs_text = ", ".join([f"{i.get('name','')}({i.get('type','')})" for i in page_data.get("inputs", [])[:10]])
    links_text = ", ".join([f"{l['text']}" for l in page_data.get("links", [])[:15] if l['text']])
    page_title = page_data.get('title', '')
    if re.search(r'parked|domain\s*for\s*sale|buy\s*this\s*domain|coming\s*soon|under\s*construction', page_title.lower()):
        print(f"[Deep] Parked domain detected: {page_title[:60]}")
        opp.deep_analysis_score = 0; opp.site_analyzed = True
        opp.profit_per_hour = "N/A (parked domain)"; opp.profit_per_day = "N/A"
        opp.profit_per_week = "N/A"; opp.profit_per_month = "N/A"; opp.profit_per_year = "N/A"
        return opp
    prompt = f"""You are a professional earnings analyst. Analyze this money-making website and provide REALISTIC estimates.

WEBSITE: {opp.title}
URL: {opp.url}
CATEGORY: {opp.category}

PAGE TITLE: {page_data.get('title','')}
PAGE TEXT (excerpt): {page_text}

VISIBLE BUTTONS: {buttons_text}
INPUT FIELDS: {inputs_text}
VISIBLE LINKS: {links_text}

CRITICAL - The LTC miner automation pattern is: Login -> Click Withdraw -> Enter Wallet -> Confirm -> Done. It's SIMPLE and fully automatable.

CRITICAL - SURVEY/GPT/TASK SITES are NOT automatable like LTC. If the site requires:
- Completing surveys (answering questions)
- Downloading apps
- Completing partner offers/offers walls
- Playing games for rewards
- Micro tasks or data entry
Then ltc_similarity_score MUST be 0-1. These sites require HUMAN WORK and cannot be automated.

TASK: Based on the page content:
1. How does this site ACTUALLY work? (describe the workflow)
2. Is it easy to automate like LTC? (login -> action -> withdraw?)
3. Does it require SURVEYS, TASKS, or OFFERS? If so, score 0-1.
4. Estimate REALISTIC earnings (be conservative - this is crypto earning, not a job):
   - Per hour (in USD)
   - Per day (in USD)
   - Per week (in USD)
   - Per month (in USD)
   - Per year (in USD)
5. What EXACT steps for a GitHub Actions Playwright bot?
6. What tools/credentials needed?
7. Rate automation similarity to LTC (0-10): 10 = exactly like LTC (login->click->withdraw)

Respond ONLY with JSON:
{{{{
  "summary": "2-3 sentence site overview",
  "workflow": "step by step how it works",
  "automation_steps": "exact steps for Playwright bot",
  "tools_needed": "tools and credentials required",
  "earn_per_hour": "$X.XX",
  "earn_per_day": "$X.XX",
  "earn_per_week": "$X.XX",
  "earn_per_month": "$X.XX",
  "earn_per_year": "$X.XX",
  "ltc_similarity_score": 0-10,
  "complexity": "Easy/Medium/Hard",
  "verdict": "GOOD TO AUTOMATE / NOT RECOMMENDED"
}}}}"""

    text = groq_generate(prompt, model=DEFAULT_GROQ_MODEL)
    parsed = False
    if text:
        try:
            match = re.search(r'\{[^}]+\}', text, re.DOTALL)
            if match:
                data = json.loads(match.group())
                opp.workflow_steps = str(data.get("workflow", ""))[:500]
                opp.automation_plan = str(data.get("automation_steps", ""))[:500]
                opp.tools_needed = str(data.get("tools_needed", ""))[:300]
                opp.profit_per_hour = str(data.get("earn_per_hour", "Unknown"))
                opp.profit_per_day = str(data.get("earn_per_day", "Unknown"))
                opp.profit_per_week = str(data.get("earn_per_week", "Unknown"))
                opp.profit_per_month = str(data.get("earn_per_month", "Unknown"))
                opp.profit_per_year = str(data.get("earn_per_year", "Unknown"))
                opp.deep_analysis_score = int(data.get("ltc_similarity_score", 0))
                opp.effort_level = str(data.get("complexity", "Unknown"))
                parsed = True
        except Exception:
            pass
    if not parsed and OPENROUTER_API_KEYS:
        print(f"[Deep] Trying OpenRouter...", end=" ", flush=True)
        text = openrouter_generate(prompt, model="openai/gpt-4o-mini")
        if text:
            try:
                match = re.search(r'\{[^}]+\}', text, re.DOTALL)
                if match:
                    data = json.loads(match.group())
                    opp.workflow_steps = str(data.get("workflow", ""))[:500]
                    opp.automation_plan = str(data.get("automation_steps", ""))[:500]
                    opp.tools_needed = str(data.get("tools_needed", ""))[:300]
                    opp.profit_per_hour = str(data.get("earn_per_hour", "Unknown"))
                    opp.profit_per_day = str(data.get("earn_per_day", "Unknown"))
                    opp.profit_per_week = str(data.get("earn_per_week", "Unknown"))
                    opp.profit_per_month = str(data.get("earn_per_month", "Unknown"))
                    opp.profit_per_year = str(data.get("earn_per_year", "Unknown"))
                    opp.deep_analysis_score = int(data.get("ltc_similarity_score", 0))
                    opp.effort_level = str(data.get("complexity", "Unknown"))
                    parsed = True
            except Exception:
                pass
    if not parsed and GEMINI_API_KEYS:
        try:
            model = genai.GenerativeModel('gemini-2.0-flash')
            resp = gemini_generate(model, prompt)
            text = resp.text.strip()
            match = re.search(r'\{[^}]+\}', text, re.DOTALL)
            if match:
                data = json.loads(match.group())
                opp.workflow_steps = str(data.get("workflow", ""))[:500]
                opp.automation_plan = str(data.get("automation_steps", ""))[:500]
                opp.tools_needed = str(data.get("tools_needed", ""))[:300]
                opp.profit_per_hour = str(data.get("earn_per_hour", "Unknown"))
                opp.profit_per_day = str(data.get("earn_per_day", "Unknown"))
                opp.profit_per_week = str(data.get("earn_per_week", "Unknown"))
                opp.profit_per_month = str(data.get("earn_per_month", "Unknown"))
                opp.profit_per_year = str(data.get("earn_per_year", "Unknown"))
                opp.deep_analysis_score = int(data.get("ltc_similarity_score", 0))
                opp.effort_level = str(data.get("complexity", "Unknown"))
                parsed = True
        except Exception as e:
            if "quota" in str(e).lower() or "429" in str(e):
                print(f"[Deep] Gemini quota exhausted")
            else:
                print(f"[Deep] Gemini error: {e}")
    if not parsed:
        return deep_analyze_rule_based(opp, page_data)
    opp.description = str(data.get("summary", opp.description))[:500]
    verdict = str(data.get("verdict", ""))
    opp.status = "confirmed" if "GOOD" in verdict.upper() else "complex"
    opp.site_analyzed = True
    captcha_info = detect_captcha_type(opp.url)
    if captcha_info.get("success") and not captcha_info.get("error"):
        captcha_info = deep_analyze_captcha(opp.url, captcha_info)
    opp.captcha_type = ", ".join(captcha_info.get("captcha_types", [])) or "NONE"
    opp.site_type = classify_site_type(opp.url, opp.title, page_text, page_data)
    no_cap_score = calculate_automation_score(captcha_info, opp.title, page_text)
    opp.no_captcha_score = no_cap_score["score"]
    ai_score = opp.deep_analysis_score
    rule_score = opp.automation_potential
    captcha_score = opp.no_captcha_score
    combined = int(ai_score * 0.5 + captcha_score * 0.3 + rule_score * 0.2)
    if opp.site_type in ("article", "directory", "telegram"):
        combined = 0
        opp.status = "complex"
    opp.deep_analysis_score = combined
    opp.automation_potential = combined
    print(f"[Deep] Score: {opp.deep_analysis_score}/10 | captcha={opp.captcha_type} | type={opp.site_type} | ${opp.profit_per_day}/day", flush=True)
    return opp

def deep_analyze_rule_based(opp: Opportunity, page_data: dict) -> Opportunity:
    body = page_data.get("body_text", "").lower()
    is_survey_gpt = any(w in body for w in [
        "survey", "surveys", "offer wall", "offerwall", "complete offers",
        "gpt", "get paid to", "paid to click", "ptc",
        "app download", "download app", "install app",
        "micro task", "microtask", "data entry", "freelance", "gig",
    ])
    is_task_based = any(w in body for w in [
        "complete tasks", "complete task", "do tasks", "task list",
        "earn points", "points for", "coin reward", "earn coins",
    ])
    is_human_work = is_survey_gpt or is_task_based
    is_mining = any(w in body for w in [
        "mining", "miner", "hash", "hashrate", "cloud mining",
        "mine crypto", "mining pool", "hash power",
    ])
    is_faucet = any(w in body for w in [
        "faucet", "claim every", "claim hourly", "claim daily",
        "free btc", "free ltc", "free doge", "free eth",
        "satoshi", "micro wallet", "dust limit",
    ])
    is_auto = any(w in body for w in [
        "auto claim", "automatic", "auto earn", "auto bot",
        "telegram bot", "auto mining", "auto withdraw",
    ])
    is_ltc_like = is_mining or is_faucet or is_auto
    has_login = any(w in body for w in ["login", "sign in", "log in", "email", "password"])
    has_withdraw = any(w in body for w in ["withdraw", "withdrawal", "claim", "send", "payout", "wallet"])
    has_register = any(w in body for w in ["register", "sign up", "create account", "get started"])
    has_earn_btn = any(w in body for w in ["click", "claim", "start", "earn"])
    pos_signals = sum([has_login, has_withdraw, has_register, has_earn_btn])
    if is_human_work and not is_ltc_like:
        opp.deep_analysis_score = 0; opp.effort_level = "Hard - requires human work"; opp.status = "complex"
        opp.workflow_steps = f"Requires manual human work: {'surveys/tasks' if is_survey_gpt else 'task completion'}. NOT automatable like LTC."
        opp.automation_plan = "Cannot automate - requires human decisions (surveys, tasks, offers, app downloads)"
        opp.tools_needed = "None - human work required, not bot-automatable"
        opp.profit_per_hour = "$0.50 - $2.00 (manual work required)"; opp.profit_per_day = "$2.00 - $10.00 (manual, not passive)"
        opp.profit_per_week = "$14.00 - $70.00 (manual work)"; opp.profit_per_month = "$60.00 - $300.00 (not recommended for automation)"
        opp.profit_per_year = "$720 - $3,600 (requires daily human effort)"
        print(f"[Deep-Rule] HUMAN WORK SITE: survey={is_survey_gpt} task={is_task_based} mining/faucet={is_ltc_like} -> score 0", flush=True)
    elif is_ltc_like and has_withdraw and pos_signals >= 2:
        opp.deep_analysis_score = 8; opp.effort_level = "Easy"; opp.status = "confirmed"
        opp.workflow_steps = "Register -> Login -> Auto-earn/Claim -> Withdraw to wallet"
        opp.automation_plan = "1. Open site 2. Login (fill email+password) 3. Click claim/earn button 4. Click withdraw 5. Enter wallet address 6. Confirm"
        opp.tools_needed = "Playwright, GitHub Actions, email+password credentials, wallet address"
        opp.profit_per_hour = "$0.01 - $0.10"; opp.profit_per_day = "$0.05 - $1.00"
        opp.profit_per_week = "$0.35 - $7.00"; opp.profit_per_month = "$1.50 - $30.00"; opp.profit_per_year = "$18 - $365"
        print(f"[Deep-Rule] LTC-MATCH: {pos_signals}/4 signals, mining={is_mining} faucet={is_faucet} -> score 8", flush=True)
    elif pos_signals >= 3 and has_withdraw:
        opp.deep_analysis_score = 4; opp.effort_level = "Medium"; opp.status = "complex"
        opp.workflow_steps = "Register -> Complete actions -> Earn -> Request withdrawal"
        opp.automation_plan = "Partial automation possible - needs investigation"
        opp.profit_per_hour = "Unknown - needs human check"; opp.profit_per_day = "Unknown - needs human check"
        opp.profit_per_week = "Unknown - needs human check"; opp.profit_per_month = "Unknown - needs human check"
        opp.profit_per_year = "Unknown - needs human check"
        print(f"[Deep-Rule] PARTIAL: {pos_signals}/4 signals, no mining/faucet -> score 4", flush=True)
    elif pos_signals >= 2:
        opp.deep_analysis_score = 2; opp.effort_level = "Medium"; opp.status = "complex"
        opp.workflow_steps = "Register -> Complete tasks -> Earn -> Request withdrawal"
        opp.automation_plan = "Investigate further - partial automation possible"
        opp.profit_per_hour = "Unknown - needs human check"; opp.profit_per_day = "Unknown - needs human check"
        opp.profit_per_week = "Unknown - needs human check"; opp.profit_per_month = "Unknown - needs human check"
        opp.profit_per_year = "Unknown - needs human check"
        print(f"[Deep-Rule] MEDIUM: {pos_signals}/4 signals, not LTC-like -> score 2", flush=True)
    else:
        opp.deep_analysis_score = 1; opp.effort_level = "Hard"; opp.status = "complex"
        opp.workflow_steps = "Unknown - site content unclear"
        opp.profit_per_hour = "Unlikely"; opp.profit_per_day = "Unlikely"; opp.profit_per_week = "Unlikely"
        opp.profit_per_month = "Unlikely"; opp.profit_per_year = "Unlikely"
        print(f"[Deep-Rule] LOW: {pos_signals}/4 signals, not suitable for automation", flush=True)
    captcha_info = detect_captcha_type(opp.url)
    if captcha_info.get("success") and not captcha_info.get("error"):
        captcha_info = deep_analyze_captcha(opp.url, captcha_info)
    opp.captcha_type = ", ".join(captcha_info.get("captcha_types", [])) or "NONE"
    opp.site_type = classify_site_type(opp.url, opp.title, body, page_data)
    no_cap_score = calculate_automation_score(captcha_info, opp.title, body)
    opp.no_captcha_score = no_cap_score["score"]
    ai_score = opp.deep_analysis_score
    rule_score = opp.automation_potential
    captcha_score = opp.no_captcha_score
    combined = int(ai_score * 0.5 + captcha_score * 0.3 + rule_score * 0.2)
    if opp.site_type in ("article", "directory", "telegram"):
        combined = 0
        opp.status = "complex"
    opp.deep_analysis_score = combined
    opp.automation_potential = combined
    opp.site_analyzed = True
    return opp
# === GOOGLE SHEETS / DOCS ===

def get_google_service(api_name, api_version, scopes):
    try:
        from google.oauth2.service_account import Credentials as SACredentials
        from googleapiclient.discovery import build
        if SERVICE_ACCOUNT_INFO:
            creds = SACredentials.from_service_account_info(SERVICE_ACCOUNT_INFO, scopes=scopes)
            print(f"[Google {api_name}] Using service account (JWT, never expires)")
            return build(api_name, api_version, credentials=creds)
    except Exception as e:
        print(f"[Google {api_name}] Service account auth error: {e}")
    try:
        if not GOOGLE_REFRESH_TOKEN:
            return None
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        creds = Credentials(token=None, client_id=GOOGLE_OAUTH_CLIENT_ID,
            client_secret=GOOGLE_OAUTH_CLIENT_SECRET,
            refresh_token=GOOGLE_REFRESH_TOKEN,
            token_uri="https://oauth2.googleapis.com/token",
            scopes=scopes)
        creds.refresh(Request())
        print(f"[Google {api_name}] Using OAuth (fallback)")
        return build(api_name, api_version, credentials=creds)
    except Exception as e:
        print(f"[Google {api_name}] Auth error: {e}")
        return None

def get_or_create_spreadsheet(service, mem):
    sheet_id = HARDCODED_SHEET_ID
    try:
        service.spreadsheets().get(spreadsheetId=sheet_id).execute()
        print(f"[Sheet] Using: https://docs.google.com/spreadsheets/d/{sheet_id}")
        mem["google_sheet_id"] = sheet_id
        return sheet_id
    except Exception as e:
        sa_email = SERVICE_ACCOUNT_INFO.get('client_email', '') if SERVICE_ACCOUNT_INFO else ''
        print(f"[Sheet] Cannot access sheet {sheet_id}: {e}")
        print(f"[Sheet] SHARE the sheet with: {sa_email} (Editor)")
        print(f"[Sheet] Or enable APIs: https://console.cloud.google.com/apis/library/sheets.googleapis.com?project=workspace-mcp-497516")
        return None

def create_run_sheet(service, sheet_id, date_str, total_new, categories, verification):
    try:
        try:
            existing = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
            for s in existing.get("sheets", []):
                if s["properties"]["title"] == date_str:
                    print(f"[Sheet] Tab '{date_str}' already exists, using it")
                    return date_str
        except:
            pass
        result = service.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": date_str}}}]}
        ).execute()
        new_sheet_id = result["replies"][0]["addSheet"]["properties"]["sheetId"]
        header_labels = [
            "ID", "Website Name", "URL", "Category", "What It Does",
            "How To Earn", "Per Hour", "Per Day", "Per Week", "Per Month", "Per Year",
            "Auto Score", "How To Automate (GitHub)", "Feasibility", "Source",
            "Found Date", "Status", "Verification",
            "Captcha Type", "Site Type", "No-Captcha Score"
        ]
        service.spreadsheets().values().update(
            spreadsheetId=sheet_id, range=f"'{date_str}'!A1",
            valueInputOption="USER_ENTERED",
            body={"values": [header_labels]}
        ).execute()
        batch_reqs = [
            {
                "repeatCell": {
                    "range": {"sheetId": new_sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                    "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
                    "fields": "userEnteredFormat.textFormat.bold"
                }
            },
            {
                "updateSheetProperties": {
                    "properties": {"sheetId": new_sheet_id, "gridProperties": {"frozenRowCount": 1}},
                    "fields": "gridProperties.frozenRowCount"
                }
            }
        ]
        service.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id,
            body={"requests": batch_reqs}
        ).execute()
        if total_new == 0:
            service.spreadsheets().values().update(
                spreadsheetId=sheet_id, range=f"'{date_str}'!A2",
                valueInputOption="USER_ENTERED",
                body={"values": [["research is still under process"]]}
            ).execute()
            print(f"[Sheet] Tab '{date_str}': 0 new - wrote 'research is still under process'")
        else:
            print(f"[Sheet] Tab '{date_str}': {total_new} new sites, header row ready")
        return date_str
    except Exception as e:
        print(f"[Sheet] Error creating run tab: {e}")
        return None

def _get_sheet_id(service, spreadsheet_id, sheet_name):
    try:
        existing = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        for s in existing.get("sheets", []):
            if s["properties"]["title"] == sheet_name:
                return s["properties"]["sheetId"]
    except:
        pass
    return 0

def load_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Opportunities"
        headers = ["ID","Title","URL","Category","Description","Profit/Hour","Profit/Day","Profit/Week","Profit/Month","Profit/Year","Effort","AutoScore","AutoReason","Source","Found Date","Tags","Status","CaptchaType","SiteType","NoCaptchaScore"]
        ws.append(headers)
        for col, w in [("A",8),("B",40),("C",50),("D",20),("E",60),("F",15),("G",15),("H",15),("I",15),("J",15),("K",12),("L",10),("M",50),("N",20),("O",20),("P",30),("Q",10),("R",15),("S",15),("T",10)]:
            ws.column_dimensions[col].width = w
        wb.save(EXCEL_FILE)
        return wb, ws
    wb = load_workbook(EXCEL_FILE)
    return wb, wb.active

def opportunity_exists(ws, url: str) -> bool:
    domain = extract_domain(url)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 2:
            existing_url = row[2] or ""
            if existing_url == url:
                return True
            if domain and extract_domain(existing_url) == domain:
                return True
    return False

def write_sheet_rows_batch(service, sheet_id, sheet_name, opps, verification):
    try:
        rows = []
        row_colors = []
        for opp in opps:
            ver = verification.get(opp.url, {})
            count = ver.get("count", 0)
            last_confirmed = ver.get("last_confirmed")
            verified_bad = ver.get("verified_bad", False)
            deep_score = ver.get("deep_score", -1)
            if verified_bad or (deep_score >= 0 and deep_score <= 1):
                verification_status = "Rejected - not automatable"
                color = {"red": 1.0, "green": 0.4, "blue": 0.4}
            elif opp.status == "confirmed" or (count >= 3 and last_confirmed):
                verification_status = "Verified"
                color = {"red": 0.8, "green": 1.0, "blue": 0.8}
            elif count >= 2:
                verification_status = f"Checking ({count}x)"
                color = {"red": 1.0, "green": 1.0, "blue": 0.6}
            elif deep_score >= 2 and deep_score <= 4:
                verification_status = "Needs review"
                color = {"red": 1.0, "green": 0.7, "blue": 0.3}
            else:
                verification_status = "Pending"
                color = {"red": 1.0, "green": 1.0, "blue": 1.0}
            rows.append([opp.id, opp.title, opp.url, opp.category, opp.description,
                opp.how_to_earn or opp.automation_reason,
                opp.profit_per_hour, opp.profit_per_day, opp.profit_per_week,
                opp.profit_per_month, opp.profit_per_year,
                opp.automation_potential,
                opp.how_to_automate, opp.feasibility,
                opp.source, opp.found_date, opp.status, verification_status,
                opp.captcha_type, opp.site_type, opp.no_captcha_score])
            row_colors.append(color)
        body = {"values": rows}
        result = service.spreadsheets().values().update(
            spreadsheetId=sheet_id,
            range=f"'{sheet_name}'!A2",
            valueInputOption="USER_ENTERED",
            body=body
        ).execute()
        written = result.get("updatedRows", len(rows))
        sheet_obj = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
        sid = None
        for s in sheet_obj.get("sheets", []):
            if s["properties"]["title"] == sheet_name:
                sid = s["properties"]["sheetId"]
                break
        if sid is not None:
            reqs = []
            for i, color in enumerate(row_colors):
                reqs.append({
                    "repeatCell": {
                        "range": {"sheetId": sid, "startRowIndex": 1 + i, "endRowIndex": 2 + i},
                        "cell": {"userEnteredFormat": {"backgroundColor": color}},
                        "fields": "userEnteredFormat.backgroundColor"
                    }
                })
            for chunk_start in range(0, len(reqs), 50):
                service.spreadsheets().batchUpdate(
                    spreadsheetId=sheet_id,
                    body={"requests": reqs[chunk_start:chunk_start + 50]}
                ).execute()
        print(f"[Sheet] {written} opps to tab '{sheet_name}'")
        return written
    except Exception as e:
        print(f"[Sheet] Batch write error: {e}")
        return 0

def write_google_doc(mem, sheet_url="", analyzed_opps=None, all_opps=None, total_new=0, categories=None):
    if not GOOGLE_DOC_ID:
        print("[Google Docs] No doc ID - skipping")
        return False
    try:
        from google.oauth2.service_account import Credentials as SACredentials
        from googleapiclient.discovery import build
        if SERVICE_ACCOUNT_INFO:
            SCOPES = ["https://www.googleapis.com/auth/documents", "https://www.googleapis.com/auth/drive"]
            creds = SACredentials.from_service_account_info(SERVICE_ACCOUNT_INFO, scopes=SCOPES)
        elif GOOGLE_REFRESH_TOKEN:
            from google.auth.transport.requests import Request
            from google.oauth2.credentials import Credentials
            creds = Credentials(token=None, client_id=GOOGLE_OAUTH_CLIENT_ID,
                client_secret=GOOGLE_OAUTH_CLIENT_SECRET,
                refresh_token=GOOGLE_REFRESH_TOKEN,
                token_uri="https://oauth2.googleapis.com/token",
                scopes=["https://www.googleapis.com/auth/documents", "https://www.googleapis.com/auth/drive"])
            creds.refresh(Request())
        else:
            print("[Google Docs] No auth method available - skipping")
            return False
        docs = build("docs", "v1", credentials=creds)
        date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        doc_id = GOOGLE_DOC_ID
        try:
            docs.documents().get(documentId=doc_id).execute()
            print(f"[Google Docs] Using permanent doc: https://docs.google.com/document/d/{doc_id}")
        except Exception as e:
            print(f"[Google Docs] Cannot access doc {doc_id}: {e}")
            return False
        all_opps = all_opps or []
        analyzed = analyzed_opps or []
        confirmed = [
            o for o in analyzed
            if o.status == "confirmed"
            and o.deep_analysis_score >= 7
            and o.site_analyzed
            and o.workflow_steps
            and len(o.workflow_steps) > 50
        ]
        content_parts = []
        content_parts.append(f"\n\n{'='*60}")
        content_parts.append(f"OPPORTUNITY HUNTER V6 - DAILY REPORT")
        content_parts.append(f"Run Date: {date_str}")
        content_parts.append(f"{'='*60}")
        content_parts.append("")
        content_parts.append("SUMMARY")
        content_parts.append("-" * 40)
        content_parts.append(f"  New sites found:    {total_new}")
        content_parts.append(f"  Total all-time:     {mem.get('total_found', 0)}")
        content_parts.append(f"  Deep-analyzed:      {len(analyzed)}")
        content_parts.append(f"  Verified automatable: {len(confirmed)}")
        content_parts.append(f"  Master Sheet:       {sheet_url if sheet_url else 'N/A'}")
        content_parts.append("")
        if categories:
            content_parts.append("CATEGORIES BREAKDOWN")
            content_parts.append("-" * 40)
            for cat, cnt in sorted(categories.items(), key=lambda x: -x[1]):
                content_parts.append(f"  {cat}: {cnt}")
            content_parts.append("")
        if all_opps:
            content_parts.append("ALL NEW SITES FOUND TODAY")
            content_parts.append("-" * 40)
            for i, opp in enumerate(all_opps, 1):
                content_parts.append(f"  {i}. {opp.title}")
                content_parts.append(f"     URL: {opp.url}")
                content_parts.append(f"     Category: {opp.category} | Score: {opp.automation_potential}/10 | Status: {opp.status}")
                content_parts.append("")
        if analyzed:
            content_parts.append("DEEP ANALYSIS RESULTS")
            content_parts.append("-" * 40)
            for opp in analyzed:
                score = opp.deep_analysis_score or opp.automation_potential
                label = "Gold" if score >= 9 else "High" if score >= 7 else "Medium" if score >= 5 else "Low"
                status = "CONFIRMED" if opp.status == "confirmed" else "REVIEW"
                content_parts.append(f"  {status}: {opp.title} ({score}/10 - {label})")
                content_parts.append(f"     URL: {opp.url}")
                content_parts.append(f"     Per Day: {opp.profit_per_day or '?'}")
                content_parts.append(f"     How: {opp.workflow_steps[:200] if opp.workflow_steps else opp.how_to_automate[:200] or 'N/A'}")
                content_parts.append("")
        if confirmed:
            content_parts.append("VERIFIED AUTOMATABLE SITES (READY TO BUILD)")
            content_parts.append("-" * 40)
            for i, opp in enumerate(confirmed, 1):
                content_parts.append(f"  SITE #{i}: {opp.title}")
                content_parts.append(f"     Score: {opp.deep_analysis_score}/10")
                content_parts.append(f"     URL: {opp.url}")
                content_parts.append(f"     Earnings: {opp.profit_per_day}/day")
                content_parts.append(f"     Workflow: {opp.workflow_steps[:300]}")
                content_parts.append("")
        text_to_append = "\n".join(content_parts)
        docs.documents().batchUpdate(
            documentId=doc_id,
            body={"requests": [
                {"insertText": {"endOfSegmentLocation": {}, "text": text_to_append}}
            ]}
        ).execute()
        print(f"[Google Docs] Daily report appended: {len(confirmed)} verified, {total_new} new sites")
        return True
    except Exception as e:
        print(f"[Google Docs] Error: {e}")
        import traceback
        traceback.print_exc()
        return False

def write_report(wb, total_new, categories, analyzed_opps=None):
    os.makedirs(REPORT_DIR, exist_ok=True)
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    path = os.path.join(REPORT_DIR, f"report_{date_str}.md")
    ws = wb.active
    total = max(0, ws.max_row - 1)
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"# Opportunity Hunter Report - {date_str}\n\n")
        f.write(f"**Total tracked (all time):** {total}  |  **New today:** {total_new}\n\n")
        if total_new == 0:
            f.write("## No new opportunities found today\n\n")
            return path
        if analyzed_opps:
            confirmed = [o for o in analyzed_opps if o.status == "confirmed"]
            f.write("## Deep Analysis Results\n\n")
            f.write(f"**Sites analyzed:** {len(analyzed_opps)}  |  **Confirmed automatable:** {len(confirmed)}\n\n")
            if confirmed:
                f.write("### CONFIRMED AUTOMATABLE (LTC-like)\n\n")
                for opp in confirmed:
                    f.write(f"#### {opp.title}\n")
                    f.write(f"- **URL:** {opp.url}\n")
                    f.write(f"- **Workflow:** {opp.workflow_steps}\n")
                    f.write(f"- **Est. Earnings:** {opp.profit_per_day}/day, {opp.profit_per_month}/month\n")
                    f.write(f"- **LTC Similarity:** {opp.deep_analysis_score}/10\n")
                    f.write(f"- **Automation Plan:** {opp.automation_plan}\n")
                    f.write(f"- **Tools:** {opp.tools_needed}\n\n")
            non_confirmed = [o for o in analyzed_opps if o.status != "confirmed"]
            if non_confirmed:
                f.write("### Complex / Not Recommended\n\n")
                for opp in non_confirmed:
                    f.write(f"- **{opp.title}** - {opp.effort_level} - {opp.workflow_steps[:100]}\n")
                f.write("\n")
        f.write("## Categories\n\n")
        for cat, cnt in sorted(categories.items(), key=lambda x: -x[1]):
            f.write(f"- **{cat}**: {cnt}\n")
        f.write("\n## Top Automatable Finds (AutoScore >= 6)\n\n")
        f.write("| Title | Auto | Category | Link |\n|---|---|---|---|\n")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 11 and row[11] and isinstance(row[11], (int,float)) and row[11] >= 6:
                f.write(f"| {str(row[1] or '')[:40]} | {row[11]}/10 | {row[3]} | {str(row[2] or '')[:50]} |\n")
    return path

def write_top_finds(ws):
    top = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 11 and row[11] and isinstance(row[11], (int,float)) and row[11] >= 7:
            top.append({"title": row[1] or "", "url": row[2] or "", "category": row[3] or "",
                "description": (row[4] or "")[:200], "automation_score": row[11],
                "automation_reason": row[12] or "", "tags": str(row[15] or "")})
    top = sorted(top, key=lambda x: -x["automation_score"])[:30]
    with open(TOP_FINDS_FILE, "w", encoding="utf-8") as f:
        json.dump({"updated": datetime.now(timezone.utc).isoformat(), "top_finds": top}, f, indent=2)
    print(f"[Top Finds] {len(top)} high-automation opportunities saved")

def send_email_notification(subject, body_html, body_text=""):
    if not GMAIL_APP_PASSWORD:
        print("[Email] GMAIL_APP_PASSWORD not set - skipping notification")
        return False
    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = GMAIL_USER
        msg['To'] = NOTIFICATION_EMAIL
        msg['Subject'] = subject
        if body_text:
            msg.attach(MIMEText(body_text, 'plain'))
        if body_html:
            msg.attach(MIMEText(body_html, 'html'))
        else:
            msg.attach(MIMEText(body_text, 'plain'))
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.send_message(msg)
        print(f"[Email] Notification sent to {NOTIFICATION_EMAIL}")
        return True
    except Exception as e:
        print(f"[Email] Failed to send: {e}")
        return False

# === MAIN ===

def main():
    try:
        with open(MEMORY_FILE, "r", encoding="utf-8", errors="replace") as f:
            mem = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        mem = {
            "runs": 0, "total_found": 0, "categories_found": {}, "last_run": None,
            "learning": [], "google_sheet_id": None, "google_doc_id": None,
            "seen_urls": [], "verification": {}
        }
    seen_urls_set = set(mem.get("seen_urls", []))
    seen_domains_set = set(mem.get("seen_domains", []))
    if seen_urls_set:
        before = len(seen_domains_set)
        for u in seen_urls_set:
            d = extract_domain(u)
            if d and 'bing.com' not in d:
                seen_domains_set.add(d)
        after = len(seen_domains_set)
        if after > before:
            print(f"[Dedup] Backfilled {after - before} new domains into seen_domains ({before} -> {after})")
    verification = mem.get("verification", {})
    mem["runs"] += 1
    wb, ws = load_excel()
    total_new, categories = 0, {}
    sheets_service = None
    sheet_id = None
    if GOOGLE_REFRESH_TOKEN:
        sheets_service = get_google_service("sheets", "v4", [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ])
        if sheets_service:
            sheet_id = get_or_create_spreadsheet(sheets_service, mem)
    else:
        print("[Sheet] No GOOGLE_REFRESH_TOKEN - Google Sheets disabled")
    pw = PlaywrightPool()
    pw.start()
    os.makedirs(DEEP_ANALYSIS_DIR, exist_ok=True)
    start_time = time.time()
    genai_available = bool(GROQ_API_KEYS) or bool(GEMINI_API_KEYS)
    new_opps = []
    all_search_results = []
    print(f"\n{'='*60}")
    print(f"PHASE 1: SEARCH & CLASSIFY ({len(ALL_QUERIES)} queries)")
    print(f"{'='*60}\n")
    for i, q in enumerate(ALL_QUERIES):
        elapsed = time.time() - start_time
        print(f"[{i+1}/{len(ALL_QUERIES)}] ({elapsed:.0f}s) {q[:55]}... ", end="", flush=True)
        items = search_all(q, pw, i)
        all_search_results.extend(items)
        new_for_query = 0
        for item in items:
            opp = classify(item, genai_scoring=genai_available)
            if opp and opp.url not in seen_urls_set and not opportunity_exists(ws, opp.url):
                domain = extract_domain(opp.url)
                if domain and domain in seen_domains_set:
                    print(f"domain dup: {domain}", end=" ", flush=True)
                    continue
                seen_urls_set.add(opp.url)
                if domain:
                    seen_domains_set.add(domain)
                ws.append([opp.id, opp.title, opp.url, opp.category, opp.description,
                    opp.profit_per_hour, opp.profit_per_day, opp.profit_per_week,
                    opp.profit_per_month, opp.profit_per_year,
                    opp.effort_level, opp.automation_potential, opp.automation_reason,
                    opp.source, opp.found_date, ",".join(opp.tags), opp.status,
                    opp.captcha_type, opp.site_type, opp.no_captcha_score])
                new_opps.append(opp)
                new_for_query += 1
                total_new += 1
                categories[opp.category] = categories.get(opp.category, 0) + 1
                mem["total_found"] += 1
        print(f"+{new_for_query}")
        mem["learning"].append({"date": datetime.now(timezone.utc).isoformat(), "query": q, "results": len(items), "new_opps": new_for_query})
        if len(mem["learning"]) > 100: mem["learning"] = mem["learning"][-100:]
        time.sleep(random.uniform(1.0, 2.5))

    # PHASE 1b: Scrapy deep crawling
    print(f"\n{'='*60}")
    print(f"PHASE 1b: DEEP CRAWLING (Scrapy)")
    print(f"{'='*60}\n")
    scrapy_results = search_scrapy("", all_search_results, pw)
    if scrapy_results:
        print(f"[Scrapy] Processing {len(scrapy_results)} deep-crawled links...")
        new_for_scrapy = 0
        for item in scrapy_results:
            opp = classify(item, genai_scoring=genai_available)
            if opp and opp.url not in seen_urls_set and not opportunity_exists(ws, opp.url):
                domain = extract_domain(opp.url)
                if domain and domain in seen_domains_set:
                    continue
                seen_urls_set.add(opp.url)
                if domain:
                    seen_domains_set.add(domain)
                ws.append([opp.id, opp.title, opp.url, opp.category, opp.description,
                    opp.profit_per_hour, opp.profit_per_day, opp.profit_per_week,
                    opp.profit_per_month, opp.profit_per_year,
                    opp.effort_level, opp.automation_potential, opp.automation_reason,
                    opp.source, opp.found_date, ",".join(opp.tags), opp.status,
                    opp.captcha_type, opp.site_type, opp.no_captcha_score])
                new_opps.append(opp)
                new_for_scrapy += 1
                total_new += 1
                categories[opp.category] = categories.get(opp.category, 0) + 1
                mem["total_found"] += 1
        print(f"[Scrapy] +{new_for_scrapy} new from deep crawl")
    else:
        print("[Scrapy] No results from deep crawling")

    wb.save(EXCEL_FILE)
    analyzed_opps = []
    if new_opps:
        top_for_analysis = sorted(new_opps, key=lambda x: -x.automation_potential)[:30]
        print(f"\n{'='*60}")
        print(f"PHASE 2: DEEP ANALYSIS ({len(top_for_analysis)} sites)")
        print(f"{'='*60}\n")
        for opp in top_for_analysis:
            opp = deep_analyze_site(pw, opp)
            analyzed_opps.append(opp)
            for row in ws.iter_rows(min_row=2):
                if row[2].value == opp.url:
                    row[5].value = opp.profit_per_hour
                    row[6].value = opp.profit_per_day
                    row[7].value = opp.profit_per_week
                    row[8].value = opp.profit_per_month
                    row[9].value = opp.profit_per_year
                    row[10].value = opp.effort_level
                    break
            time.sleep(random.uniform(1.0, 2.0))
        print(f"\n[Deep] Analysis complete. Confirmed: {len([o for o in analyzed_opps if o.status == 'confirmed'])}/{len(analyzed_opps)}")
    pw.close()
    wb.save(EXCEL_FILE)
    run_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    mem["last_run"] = datetime.now(timezone.utc).isoformat()
    for cat, cnt in categories.items():
        mem["categories_found"][cat] = mem["categories_found"].get(cat, 0) + cnt
    mem["seen_urls"] = sorted(seen_urls_set)
    mem["seen_domains"] = sorted(seen_domains_set)
    for opp in analyzed_opps:
        if opp.url not in verification:
            verification[opp.url] = {"count": 0, "first_seen": run_date, "last_confirmed": None}
        verification[opp.url]["count"] += 1
        if opp.status == "confirmed" and opp.deep_analysis_score >= 7:
            verification[opp.url]["last_confirmed"] = run_date
        if opp.site_analyzed:
            verification[opp.url]["deep_score"] = opp.deep_analysis_score
            if opp.deep_analysis_score <= 1:
                verification[opp.url]["verified_bad"] = True
            else:
                verification[opp.url]["verified_bad"] = False
    mem["verification"] = verification
    with open(MEMORY_FILE, "w", encoding="utf-8") as f:
        json.dump(mem, f, indent=2)
    if sheets_service and sheet_id:
        run_sheet_name = create_run_sheet(sheets_service, sheet_id, run_date, total_new, categories, verification)
        if run_sheet_name and new_opps:
            write_sheet_rows_batch(sheets_service, sheet_id, run_sheet_name, new_opps, verification)
    elif sheet_id:
        print(f"[Sheet] No new opportunities found - no daily sheet created")
    write_top_finds(ws)
    report_path = write_report(wb, total_new, categories, analyzed_opps)
    elapsed = time.time() - start_time
    sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}" if sheet_id else ""
    print(f"\nRun #{mem['runs']}: +{total_new} new, {mem['total_found']} total in {elapsed:.0f}s")
    if sheet_url:
        print(f"Master Sheet: {sheet_url}")
    strict_confirmed = [
        o for o in (analyzed_opps or [])
        if o.status == "confirmed"
        and o.deep_analysis_score >= 7
        and o.site_analyzed
        and o.workflow_steps
        and len(o.workflow_steps) > 50
    ]
    write_google_doc(mem, sheet_url, analyzed_opps, new_opps, total_new, categories)
    with open(MEMORY_FILE, "w", encoding="utf-8") as f:
        json.dump(mem, f, indent=2)
    run_num = mem['runs']
    date_today = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    subject = f"Opportunity Hunter V6 - Run #{run_num} Complete ({date_today})"
    body_html = f"""
    <html><body style="font-family: Arial, sans-serif; max-width: 700px; margin: 20px;">
    <h2 style="color: #2563eb;">Opportunity Hunter V6 - Daily Report</h2>
    <p><strong>Run #{run_num}</strong> | <strong>{date_today}</strong></p>
    <hr style="border: 1px solid #e5e7eb;">
    <table style="width: 100%; border-collapse: collapse; margin: 15px 0;">
        <tr style="background: #f3f4f6;">
            <th style="padding: 8px; text-align: left; border: 1px solid #d1d5db;">Metric</th>
            <th style="padding: 8px; text-align: right; border: 1px solid #d1d5db;">Value</th>
        </tr>
        <tr>
            <td style="padding: 8px; border: 1px solid #d1d5db;">New Sites Found</td>
            <td style="padding: 8px; text-align: right; border: 1px solid #d1d5db;">{total_new}</td>
        </tr>
        <tr style="background: #f9fafb;">
            <td style="padding: 8px; border: 1px solid #d1d5db;">Deep-Analyzed</td>
            <td style="padding: 8px; text-align: right; border: 1px solid #d1d5db;">{len(analyzed_opps) if analyzed_opps else 0}</td>
        </tr>
        <tr>
            <td style="padding: 8px; border: 1px solid #d1d5db;">Verified (score >=7)</td>
            <td style="padding: 8px; text-align: right; border: 1px solid #d1d5db;">{len(strict_confirmed)}</td>
        </tr>
        <tr style="background: #f9fafb;">
            <td style="padding: 8px; border: 1px solid #d1d5db;">Total All-Time</td>
            <td style="padding: 8px; text-align: right; border: 1px solid #d1d5db;">{mem['total_found']}</td>
        </tr>
        <tr>
            <td style="padding: 8px; border: 1px solid #d1d5db;">Run Time</td>
            <td style="padding: 8px; text-align: right; border: 1px solid #d1d5db;">{elapsed:.0f}s</td>
        </tr>
    </table>"""
    if strict_confirmed:
        body_html += """
    <h3 style="color: #059669;">Verified Automatable Sites</h3>
    <table style="width: 100%; border-collapse: collapse; margin: 15px 0;">
        <tr style="background: #059669; color: white;">
            <th style="padding: 8px; text-align: left; border: 1px solid #d1d5db;">Site</th>
            <th style="padding: 8px; text-align: center; border: 1px solid #d1d5db;">Score</th>
            <th style="padding: 8px; text-align: right; border: 1px solid #d1d5db;">Per Day</th>
        </tr>"""
        for opp in strict_confirmed:
            body_html += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #d1d5db;">{opp.title}</td>
            <td style="padding: 8px; text-align: center; border: 1px solid #d1d5db;">{opp.deep_analysis_score}/10</td>
            <td style="padding: 8px; text-align: right; border: 1px solid #d1d5db;">{opp.profit_per_day or '?'}</td>
        </tr>"""
        body_html += "</table>"
    body_html += """<hr style="border: 1px solid #e5e7eb;">"""
    if sheet_url:
        body_html += f'<p><a href="{sheet_url}">View Google Sheet</a></p>'
    if GOOGLE_DOC_ID:
        body_html += f'<p><a href="https://docs.google.com/document/d/{GOOGLE_DOC_ID}">View Verified Sites Doc</a></p>'
    body_html += """
    <p style="color: #6b7280; font-size: 12px;">This is an automated report from Opportunity Hunter V6 running on GitHub Actions.</p>
    </body></html>"""
    body_text = f"""Opportunity Hunter V6 - Run #{run_num} Complete
Date: {date_today}
Results:
  New Sites Found: {total_new}
  Deep-Analyzed: {len(analyzed_opps) if analyzed_opps else 0}
  Verified (score >=7): {len(strict_confirmed)}
  Total All-Time: {mem['total_found']}
  Run Time: {elapsed:.0f}s
Sheet: {sheet_url if sheet_url else 'N/A'}
Doc: https://docs.google.com/document/d/{GOOGLE_DOC_ID if GOOGLE_DOC_ID else 'N/A'}"""
    if strict_confirmed:
        body_text += f"\nVerified Sites ({len(strict_confirmed)}):\n"
        for opp in strict_confirmed:
            body_text += f"  - {opp.title} ({opp.deep_analysis_score}/10): {opp.profit_per_day or '?'}/day\n"
    send_email_notification(subject, body_html, body_text)

    # === WEBHOOK NOTIFICATION ===
    try:
        webhook_url = os.environ.get("WEBHOOK_URL", "https://notify.opencode.co/webhook/opportunity-hunt")
        if webhook_url:
            webhook_payload = {
                "event": "run_completed",
                "run_number": run_num,
                "date": date_today,
                "total_new": total_new,
                "total_all_time": mem['total_found'],
                "deep_analyzed": len(analyzed_opps) if analyzed_opps else 0,
                "verified_count": len(strict_confirmed),
                "elapsed_seconds": int(elapsed),
                "sheet_url": sheet_url if sheet_id else "",
                "doc_url": f"https://docs.google.com/document/d/{GOOGLE_DOC_ID}" if GOOGLE_DOC_ID else "",
                "verified_sites": [
                    {"title": opp.title, "url": opp.url, "score": opp.deep_analysis_score,
                     "profit_per_day": opp.profit_per_day or ""}
                    for opp in strict_confirmed
                ] if strict_confirmed else []
            }
            r = requests.post(webhook_url, json=webhook_payload, timeout=15,
                headers={"User-Agent": "OpportunityHunterV6/1.0"})
            if r.status_code == 200:
                print(f"[Webhook] Sent to {webhook_url}")
            else:
                print(f"[Webhook] HTTP {r.status_code} from {webhook_url}")
    except Exception as e:
        print(f"[Webhook] Error: {e}")
    print(f"[{datetime.now(timezone.utc).isoformat()}] Done!")

if __name__ == "__main__":
    main()
