import os, json, time, re, hashlib, sys, urllib.parse, html as html_mod, random
from datetime import datetime, timezone
from typing import Optional
from dataclasses import dataclass, field

os.system("pip install requests openpyxl google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client playwright google-generativeai 2>/dev/null")
os.system("playwright install chromium 2>/dev/null")

import requests
from openpyxl import Workbook, load_workbook

FIRECRAWL_API_KEY = os.environ.get("FIRECRAWL_API_KEY", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GOOGLE_OAUTH_CLIENT_ID = os.environ.get("GOOGLE_OAUTH_CLIENT_ID", "")
GOOGLE_OAUTH_CLIENT_SECRET = os.environ.get("GOOGLE_OAUTH_CLIENT_SECRET", "")
GOOGLE_REFRESH_TOKEN = os.environ.get("GOOGLE_REFRESH_TOKEN", "")

EXCEL_FILE = "opportunities.xlsx"
MEMORY_FILE = "bot_memory.json"
REPORT_DIR = "reports"
TOP_FINDS_FILE = "top_automatable_finds.json"

SHEET_COLUMNS = [
    "ID", "Website Name", "URL", "Category", "What It Does",
    "How To Earn", "Per Hour", "Per Day", "Per Week", "Per Month", "Per Year",
    "Auto Score", "How To Automate (GitHub)", "Feasibility", "Source",
    "Found Date", "Status"
]

@dataclass
class Opportunity:
    id: str = ""
    title: str = ""
    url: str = ""
    category: str = ""
    description: str = ""
    profit_per_day: str = ""
    profit_per_week: str = ""
    profit_per_month: str = ""
    effort_level: str = ""
    automation_potential: int = 0
    automation_reason: str = ""
    source: str = ""
    found_date: str = ""
    tags: list = field(default_factory=list)
    status: str = "new"
    how_to_earn: str = ""
    how_to_automate: str = ""
    feasibility: str = ""

def get_google_service(api_name, api_version, scopes):
    if not GOOGLE_REFRESH_TOKEN:
        return None
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        creds = Credentials(token=None, client_id=GOOGLE_OAUTH_CLIENT_ID,
            client_secret=GOOGLE_OAUTH_CLIENT_SECRET,
            refresh_token=GOOGLE_REFRESH_TOKEN,
            token_uri="https://oauth2.googleapis.com/token",
            scopes=scopes)
        creds.refresh(Request())
        return build(api_name, api_version, credentials=creds)
    except Exception as e:
        print(f"[Google {api_name}] Auth error: {e}")
        return None

def setup_google_sheet(service, mem):
    sheet_id = mem.get("google_sheet_id")
    if sheet_id:
        try:
            service.spreadsheets().get(spreadsheetId=sheet_id).execute()
            print(f"[Sheet] Using: https://docs.google.com/spreadsheets/d/{sheet_id}")
            return sheet_id
        except:
            print("[Sheet] Stored ID stale, creating new...")
            mem.pop("google_sheet_id", None)
    sheet = service.spreadsheets().create(body={
        "properties": {"title": "Opportunity Hunter - Master Findings"}
    }).execute()
    sheet_id = sheet["spreadsheetId"]
    service.spreadsheets().values().update(
        spreadsheetId=sheet_id, range="A1:Q1",
        valueInputOption="USER_ENTERED",
        body={"values": [SHEET_COLUMNS]}
    ).execute()
    service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body={"requests": [{
        "repeatCell": {
            "range": {"startRowIndex": 0, "endRowIndex": 1},
            "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
            "fields": "userEnteredFormat.textFormat.bold"
        }
    }]}).execute()
    mem["google_sheet_id"] = sheet_id
    print(f"[Sheet] Created: https://docs.google.com/spreadsheets/d/{sheet_id}")
    return sheet_id

def append_google_sheet_row(service, sheet_id, opp):
    row = [opp.id, opp.title, opp.url, opp.category, opp.description,
        opp.how_to_earn or opp.automation_reason,
        opp.profit_per_day, opp.profit_per_week, opp.profit_per_month,
        "", "", opp.automation_potential,
        opp.how_to_automate, opp.feasibility,
        opp.source, opp.found_date, opp.status]
    try:
        service.spreadsheets().values().append(
            spreadsheetId=sheet_id, range="A:Q",
            valueInputOption="USER_ENTERED",
            insertDataOption="INSERT_ROWS",
            body={"values": [row]}
        ).execute()
        print(f"[Sheet] Appended: {opp.title[:40]}")
    except Exception as e:
        print(f"[Sheet] Append error: {e}")

def load_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Opportunities"
        headers = ["ID","Title","URL","Category","Description","Profit/Day","Profit/Week","Profit/Month","Effort","AutoScore","AutoReason","Source","Found Date","Tags","Status"]
        ws.append(headers)
        for col, w in [("A",8),("B",40),("C",50),("D",20),("E",60),("F",15),("G",15),("H",15),("I",12),("J",10),("K",50),("L",20),("M",20),("N",30),("O",10)]:
            ws.column_dimensions[col].width = w
        wb.save(EXCEL_FILE)
        return wb, ws
    wb = load_workbook(EXCEL_FILE)
    return wb, wb.active

def opportunity_exists(ws, url: str) -> bool:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 2 and row[2] == url:
            return True
    return False

def extract_ddg_url(u: str) -> str:
    if "duckduckgo.com/l/" in u:
        parsed = urllib.parse.urlparse(u)
        qs = urllib.parse.parse_qs(parsed.query)
        return qs.get("uddg", [u])[0]
    return u

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
]

def search_bing_html(query: str, ua_idx: int = 0) -> list:
    try:
        resp = requests.get("https://www.bing.com/search", params={"q": query, "count": 10},
            headers={"User-Agent": USER_AGENTS[ua_idx % len(USER_AGENTS)], "Accept-Language": "en-US,en;q=0.9"},
            timeout=10)
        links = re.findall(r'<a[^>]+href="(https?://[^"]+)"[^>]*><h2>(.*?)</h2>', resp.text, re.DOTALL)
        if not links:
            links = re.findall(r'<h2[^>]*><a[^>]+href="(https?://[^"]+)"[^>]*>(.*?)</a></h2>', resp.text, re.DOTALL)
        snippets = re.findall(r'<p[^>]*>(.*?)</p>', resp.text, re.DOTALL)
        results = []
        for i, (u, t) in enumerate(links[:10]):
            desc = html_mod.unescape(re.sub(r'<[^>]+>', '', snippets[i] if i < len(snippets) else "")).strip()[:300] if i < len(snippets) else ""
            results.append({"title": html_mod.unescape(re.sub(r'<[^>]+>', '', t)).strip()[:200], "url": u, "description": desc})
        return results
    except Exception as e:
        print(f"[Bing HTML] Error: {e}")
        return []

def search_ddg(query: str, ua_idx: int = 0) -> list:
    try:
        resp = requests.get("https://html.duckduckgo.com/html/", params={"q": query},
            headers={"User-Agent": USER_AGENTS[ua_idx % len(USER_AGENTS)]}, timeout=8)
        links = re.findall(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', resp.text, re.DOTALL)
        snippets = re.findall(r'<a[^>]+class="result__snippet"[^>]*>(.*?)</a>', resp.text, re.DOTALL)
        if not links:
            return []
        results = []
        for i, (u, t) in enumerate(links[:10]):
            desc = html_mod.unescape(re.sub(r'<[^>]+>', '', snippets[i] if i < len(snippets) else "")).strip()[:300] if i < len(snippets) else ""
            url = extract_ddg_url(u)
            if url:
                results.append({"title": html_mod.unescape(re.sub(r'<[^>]+>', '', t)).strip()[:200], "url": url, "description": desc})
        return results
    except:
        return []

def search_startpage(query: str, ua_idx: int = 0) -> list:
    try:
        resp = requests.post("https://www.startpage.com/sp/search",
            data={"query": query, "language": "en", "cat": "web", "page": 1},
            headers={"User-Agent": USER_AGENTS[ua_idx % len(USER_AGENTS)]}, timeout=10)
        if resp.status_code == 200:
            links = re.findall(r'<a[^>]+class="w-gl__result-title"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', resp.text, re.DOTALL)
            descs = re.findall(r'<p class="w-gl__description[^>]*>(.*?)</p>', resp.text, re.DOTALL)
            results = []
            for i, (u, t) in enumerate(links[:10]):
                desc = html_mod.unescape(re.sub(r'<[^>]+>', '', descs[i] if i < len(descs) else "")).strip()[:300]
                results.append({"title": html_mod.unescape(re.sub(r'<[^>]+>', '', t)).strip()[:200], "url": u, "description": desc})
            return results
        return []
    except:
        return []

class PlaywrightPool:
    def __init__(self):
        self.browser = None
        self._pw = None
    def start(self):
        from playwright.sync_api import sync_playwright
        self._pw = sync_playwright()
        p = self._pw.start()
        self.browser = p.chromium.launch(headless=True, args=["--no-sandbox","--disable-setuid-sandbox","--window-size=1920,1080"])
    def search(self, query: str, q_idx: int = 0) -> list:
        if not self.browser:
            self.start()
        try:
            ctx = self.browser.new_context(
                user_agent=USER_AGENTS[q_idx % len(USER_AGENTS)],
                viewport={"width": 1920, "height": 1080},
                locale="en-US",
            )
            page = ctx.new_page()
            page.goto(f"https://www.bing.com/search?q={urllib.parse.quote(query)}&count=10", timeout=20000)
            page.wait_for_timeout(2000)
            if "captcha" in page.content().lower():
                ctx.close()
                return []
            page.evaluate("window.scrollBy(0, 400)")
            page.wait_for_timeout(500)
            items = page.query_selector_all("li.b_algo")
            results = []
            for el in items[:12]:
                try:
                    h2 = el.query_selector("h2")
                    a = h2.query_selector("a[href^='http']") if h2 else None
                    if not a: continue
                    url = a.get_attribute("href") or ""
                    title = h2.inner_text().strip()
                    desc_el = el.query_selector("div.b_caption p, div.b_snippet")
                    desc = desc_el.inner_text()[:400] if desc_el else ""
                    if url and title:
                        results.append({"title": title, "url": url, "description": desc.strip()})
                except: continue
            ctx.close()
            return results
        except Exception as e:
            print(f"[Bing PW] Error: {e}")
            return []
    def close(self):
        if self.browser:
            self.browser.close()
        if self._pw:
            try: self._pw.__exit__(None, None, None)
            except: pass

def search_all(query: str, pw: PlaywrightPool, q_idx: int) -> list:
    seen = set()
    results = []
    pw_results = pw.search(query, q_idx)
    print(f"[Bing PW] {len(pw_results)} results", end=" ", flush=True)
    for r in pw_results:
        if r["url"] and r["url"] not in seen:
            seen.add(r["url"])
            results.append(r)
    time.sleep(random.uniform(0.3, 0.8))
    html_results = search_bing_html(query, q_idx)
    print(f"[Bing HTML] {len(html_results)}", end=" ", flush=True)
    for r in html_results:
        if r["url"] and r["url"] not in seen:
            seen.add(r["url"])
            results.append(r)
    if len(results) < 3:
        time.sleep(random.uniform(0.3, 0.8))
        ddg_results = search_ddg(query, q_idx)
        print(f"[DDG] {len(ddg_results)}", end=" ", flush=True)
        for r in ddg_results:
            if r["url"] and r["url"] not in seen:
                seen.add(r["url"])
                results.append(r)
    if len(results) < 2:
        time.sleep(random.uniform(0.3, 0.8))
        sp_results = search_startpage(query, q_idx)
        print(f"[SP] {len(sp_results)}", end=" ", flush=True)
        for r in sp_results:
            if r["url"] and r["url"] not in seen:
                seen.add(r["url"])
                results.append(r)
    print(f"=> {len(results)} unique", flush=True)
    return results

def score_automation_gemini(title: str, desc: str, url: str) -> tuple:
    if not GEMINI_API_KEY:
        return rule_score_automation(title, desc, url)
    try:
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-2.0-flash')
        prompt = f"""Rate this opportunity's AUTOMATION POTENTIAL (0-10).
High (7-10): fully automatable with a GitHub bot - auto-click, auto-claim, auto-mining, auto-faucet, auto-task scripts. No human needed.
Medium (4-6): partially automatable - needs occasional captchas or approvals.
Low (0-3): requires continuous human work - typing, reading, manual trading.

Title: {title[:200]}
Description: {desc[:300]}
URL: {url[:200]}

Respond ONLY with JSON: {{"score": N, "reason": "short reason", "how_to_earn": "how to earn from this", "how_to_automate": "how to automate with GitHub Actions"}}"""
        resp = model.generate_content(prompt)
        text = resp.text.strip()
        match = re.search(r'\{[^}]+\}', text)
        if match:
            data = json.loads(match.group())
            score = max(0, min(10, int(data.get("score", 5))))
            return score, str(data.get("reason", ""))[:200], str(data.get("how_to_earn", ""))[:300], str(data.get("how_to_automate", ""))[:300]
        return 5, "Parse error", "", ""
    except Exception as e:
        print(f"[Gemini] Error: {e}")
        return rule_score_automation(title, desc, url)

def rule_score_automation(title: str, desc: str, url: str) -> tuple:
    c = f"{title} {desc} {url}".lower()
    if any(s in c for s in ["dictionary","meaning","definition","wikipedia","encyclopedia","tutorial","course","educational","academic"]):
        return 1, "Educational content", "", ""
    bonus, signals = 0, 0
    for words, pts in [(["mining","miner","hash","cloud mining","free btc","free eth","free ltc","free doge"], 3),
                       (["faucet","claim","withdraw","crypto faucet"], 3),
                       (["auto","bot","telegram bot","auto claim","auto bot"], 2),
                       (["click","ptc","paid-to-click","earn per"], 2),
                       (["passive income","staking","stake","passive"], 2),
                       (["refer","affiliate","earn crypto","reward","bonus","cash"], 1)]:
        if any(w in c for w in words):
            bonus += pts
            signals += 1
    if "survey" in c or "data entry" in c:
        bonus -= 2
    score = min(10, max(0, 5 + bonus // max(1, signals // 2 if signals else 1)))
    if signals == 0:
        return 0, "Low: no automation signals", "", ""
    how_to_earn = "Visit website, complete tasks/claims, withdraw earnings to wallet"
    how_to_automate = "Use Playwright browser automation on GitHub Actions to login, claim, and withdraw on schedule"
    if score >= 7:
        return score, "High automation: bot-friendly signals", how_to_earn, how_to_automate
    elif score >= 4:
        return score, "Medium: partially automatable", how_to_earn, how_to_automate
    return score, "Low: human interaction needed", how_to_earn, how_to_automate

CLASSIFY_KEYWORDS = [
    (["faucet","free crypto","claim","btc faucet","eth faucet","crypto faucet","bitcoin faucet"], "Crypto Faucet", ["faucet","crypto","free"]),
    (["airdrop","token distribution","free token","claim airdrop","crypto airdrop"], "Airdrop", ["airdrop","crypto","free"]),
    (["ptc","paid-to-click","bux","click ads","get paid to","paidtoclick","earn per click"], "PTC / GPT", ["ptc","gpt","click"]),
    (["survey","paid survey","market research","paid surveys"], "Paid Surveys", ["survey","research"]),
    (["cashback","cash back","rebate","shopping","cashback site"], "Cashback", ["cashback","shopping"]),
    (["affiliate","referral","refer","affiliate program"], "Affiliate", ["affiliate","referral"]),
    (["stake","staking","defi","yield","lend","apr","liquidity","pool"], "DeFi / Staking", ["defi","staking","yield"]),
    (["play to earn","p2e","gamefi","nft game","play-to-earn","crypto game"], "Play-to-Earn", ["p2e","gaming","nft"]),
    (["mining","cloud mining","hash","mine","miner","crypto mining","ltc miner","bitcoin miner","free mining","mining pool"], "Mining", ["mining","crypto"]),
    (["trading bot","auto trade","signal","copy trade","trading platform","grid trading"], "Trading", ["trading","bot","automation"]),
    (["micro task","microtask","data entry","freelance","micro job","gig"], "Micro Tasks", ["micro-task","freelance"]),
    (["browser automation","auto earn","auto claim","auto bot","automation bot","auto click","auto faucet"], "Automation Bot", ["automation","bot"]),
    (["earn crypto","free crypto","get crypto","crypto earn","crypto reward","crypto bonus"], "Crypto Earnings", ["crypto","earn"]),
    (["ltc","litecoin","dogecoin","doge coin","doge mining","ltc mining","free ltc","free doge"], "Altcoin Mining", ["mining","ltc","doge"]),
]

def classify(item: dict, genai_scoring: bool = True) -> Optional[Opportunity]:
    t, u, d = item.get("title",""), item.get("url",""), item.get("description","")
    c = f"{t} {d}".lower()
    for keywords, cat, tags in CLASSIFY_KEYWORDS:
        if any(k in c for k in keywords):
            if genai_scoring:
                auto_score, auto_reason, how_to_earn, how_to_automate = score_automation_gemini(t, d, u)
            else:
                auto_score, auto_reason, how_to_earn, how_to_automate = rule_score_automation(t, d, u)
            fea = "Easy" if auto_score >= 7 else "Medium" if auto_score >= 4 else "Hard"
            return Opportunity(id=hashlib.md5(f"{t}{u}{time.time()}".encode()).hexdigest()[:12],
                title=t[:120], url=u[:300], category=cat, description=d[:500],
                profit_per_day="Unknown", profit_per_week="Unknown", profit_per_month="Unknown",
                effort_level="Unknown", automation_potential=auto_score, automation_reason=auto_reason,
                how_to_earn=how_to_earn, how_to_automate=how_to_automate, feasibility=fea,
                source="web_search", found_date=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                tags=tags, status="new")
    if any(k in c for k in ["earn","money","income","profit","passive","free","reward","bonus","cash","crypto","bitcoin"]):
        if genai_scoring:
            auto_score, auto_reason, how_to_earn, how_to_automate = score_automation_gemini(t, d, u)
        else:
            auto_score, auto_reason, how_to_earn, how_to_automate = rule_score_automation(t, d, u)
        fea = "Easy" if auto_score >= 7 else "Medium" if auto_score >= 4 else "Hard"
        return Opportunity(id=hashlib.md5(f"{t}{u}{time.time()}".encode()).hexdigest()[:12],
            title=t[:120], url=u[:300], category="General", description=d[:500],
            profit_per_day="Unknown", profit_per_week="Unknown", profit_per_month="Unknown",
            effort_level="Unknown", automation_potential=auto_score, automation_reason=auto_reason,
            how_to_earn=how_to_earn, how_to_automate=how_to_automate, feasibility=fea,
            source="web_search", found_date=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            tags=["money","earn"], status="new")
    return None

QUERIES = [
    "free ltc mining sites 2026 no deposit instant withdrawal",
    "free dogecoin mining sites 2026 no deposit",
    "free bitcoin mining cloud mining 2026 no deposit withdraw",
    "free ethereum mining sites 2026 no investment",
    "crypto mining faucet earn free btc eth ltc doge 2026",
    "best crypto faucets 2026 free bitcoin ethereum litecoin",
    "auto claim crypto faucet bot 2026",
    "paid to click sites that pay instantly 2026 free registration",
    "best GPT sites earn money free 2026 auto earn",
    "new crypto airdrops 2026 free tokens claim",
    "solana airdrop 2026 claim free",
    "telegram bot airdrop claim 2026",
    "passive income crypto staking defi 2026 no minimum",
    "free crypto staking rewards 2026",
    "browser automation earn crypto free 2026",
    "telegram bot earn crypto 2026 free automated",
    "auto trading crypto bot free 2026",
    "free crypto arbitrage bot 2026",
    "play to earn crypto games 2026 free no investment",
    "micro task sites pay crypto 2026",
    "best cashback apps 2026 free money crypto",
    "affiliate programs crypto 2026 high paying free",
    "earn free crypto no deposit 2026 withdraw instantly",
    "free litecoin mining pool 2026 no deposit required",
    "doge coin faucet free claim every hour 2026",
    "btc mining telegram bot free 2026",
    "automatic crypto earning platform 2026 no investment",
    "free bitcoin earning sites 2026 withdraw to wallet",
    "cloud mining free trial 2026 no deposit btc",
    "web3 earn crypto free 2026 browser mining",
    "free crypto signals telegram 2026 copy trade",
    "defi yield farming 2026 no minimum deposit",
    "passive crypto income 2026 set and forget",
    "telegram mining bot free withdrawal 2026",
    "faucet pay crypto instant 2026 free claim",
]

def write_report(wb, total_new, categories):
    os.makedirs(REPORT_DIR, exist_ok=True)
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    path = os.path.join(REPORT_DIR, f"report_{date_str}.md")
    ws = wb.active
    total = max(0, ws.max_row - 1)
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"# Opportunity Hunter Report - {date_str}\n\n")
        f.write(f"**Total tracked:** {total}  |  **New today:** {total_new}\n\n")
        f.write("## Categories\n\n")
        for cat, cnt in sorted(categories.items(), key=lambda x: -x[1]):
            f.write(f"- **{cat}**: {cnt}\n")
        f.write("\n## Top Automatable Finds (AutoScore >= 6)\n\n")
        f.write("| Title | Auto | Category | Link |\n|---|---|---|---|\n")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 9 and row[9] and isinstance(row[9], (int,float)) and row[9] >= 6:
                f.write(f"| {str(row[1] or '')[:40]} | {row[9]}/10 | {row[3]} | {str(row[2] or '')[:50]} |\n")
        f.write("\n## Latest 20\n\n| Title | Category | AutoScore | Link |\n|---|---|---|---|\n")
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if i >= 20: break
            if len(row) > 9:
                f.write(f"| {str(row[1] or '')[:40]} | {row[3]} | {row[9] or '?'}/10 | {str(row[2] or '')[:50]} |\n")
    return path

def write_top_finds(ws):
    top = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 9 and row[9] and isinstance(row[9], (int,float)) and row[9] >= 7:
            top.append({"title": row[1] or "", "url": row[2] or "", "category": row[3] or "",
                "description": (row[4] or "")[:200], "automation_score": row[9],
                "automation_reason": row[10] or "", "tags": str(row[13] or "")})
    top = sorted(top, key=lambda x: -x["automation_score"])[:30]
    with open(TOP_FINDS_FILE, "w", encoding="utf-8") as f:
        json.dump({"updated": datetime.now(timezone.utc).isoformat(), "top_finds": top}, f, indent=2)
    print(f"[Top Finds] {len(top)} high-automation opportunities saved")

def write_google_doc(report_path, sheet_url=""):
    if not GOOGLE_REFRESH_TOKEN:
        print("[Google Docs] No token - skipping")
        return False
    try:
        from google.auth.transport.requests import Request as GoogleRequest
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        creds = Credentials(token=None, client_id=GOOGLE_OAUTH_CLIENT_ID,
            client_secret=GOOGLE_OAUTH_CLIENT_SECRET,
            refresh_token=GOOGLE_REFRESH_TOKEN,
            token_uri="https://oauth2.googleapis.com/token")
        creds.refresh(GoogleRequest())
        docs = build("docs", "v1", credentials=creds)
        date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        doc = docs.documents().create(body={"title": f"Opportunity Hunter Report - {date_str}"}).execute()
        doc_id = doc["documentId"]
        with open(report_path, encoding="utf-8") as f:
            content = f.read()
        clean_lines = []
        for line in content.split('\n'):
            line = re.sub(r'^#{1,6}\s+', '', line)
            line = line.replace('**', '').replace('*', '').replace('__', '').replace('_', '')
            if re.match(r'^\|[\s\-:]+\|$', line): continue
            if line.startswith('|') and '|' in line[1:]:
                cells = [c.strip() for c in line.split('|')[1:-1]]
                line = ' | '.join(cells)
            clean_lines.append(line)

        if sheet_url:
            clean_text = f"Master Sheet: {sheet_url}\n\n" + '\n'.join(clean_lines)
        else:
            clean_text = '\n'.join(clean_lines)

        docs.documents().batchUpdate(documentId=doc_id,
            body={"requests": [{"insertText": {"endOfSegmentLocation": {}, "text": clean_text}}]}).execute()
        print(f"[Google Docs] Created: https://docs.google.com/document/d/{doc_id}")
        return True
    except Exception as e:
        print(f"[Google Docs] Error: {e}")
        return False

def main():
    mem = json.load(open(MEMORY_FILE)) if os.path.exists(MEMORY_FILE) else {
        "runs": 0, "total_found": 0, "categories_found": {}, "last_run": None,
        "learning": [], "google_sheet_id": None
    }
    mem["runs"] += 1
    wb, ws = load_excel()
    total_new, categories = 0, {}

    sheets_service = None
    sheet_id = None
    if GOOGLE_REFRESH_TOKEN:
        sheets_service = get_google_service("sheets", "v4", [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ])
        if sheets_service:
            sheet_id = setup_google_sheet(sheets_service, mem)
    else:
        print("[Sheet] No GOOGLE_REFRESH_TOKEN - Google Sheets disabled")

    pw = PlaywrightPool()
    pw.start()
    start_time = time.time()

    genai_available = bool(GEMINI_API_KEY)

    for i, q in enumerate(QUERIES):
        elapsed = time.time() - start_time
        print(f"[{i+1}/{len(QUERIES)}] ({elapsed:.0f}s) {q[:55]}... ", end="", flush=True)
        items = search_all(q, pw, i)
        new_for_query = 0
        for item in items:
            opp = classify(item, genai_scoring=genai_available)
            if opp and not opportunity_exists(ws, opp.url):
                ws.append([opp.id, opp.title, opp.url, opp.category, opp.description,
                    opp.profit_per_day, opp.profit_per_week, opp.profit_per_month,
                    opp.effort_level, opp.automation_potential, opp.automation_reason,
                    opp.source, opp.found_date, ",".join(opp.tags), opp.status])
                if sheets_service and sheet_id:
                    append_google_sheet_row(sheets_service, sheet_id, opp)
                new_for_query += 1
                total_new += 1
                categories[opp.category] = categories.get(opp.category, 0) + 1
                mem["total_found"] += 1
        print(f"+{new_for_query}")
        mem["learning"].append({"date": datetime.now(timezone.utc).isoformat(), "query": q, "results": len(items), "new_opps": new_for_query})
        if len(mem["learning"]) > 100: mem["learning"] = mem["learning"][-100:]
        time.sleep(random.uniform(1.0, 2.5))

    pw.close()
    wb.save(EXCEL_FILE)
    mem["last_run"] = datetime.now(timezone.utc).isoformat()
    for cat, cnt in categories.items():
        mem["categories_found"][cat] = mem["categories_found"].get(cat, 0) + cnt
    json.dump(mem, open(MEMORY_FILE, "w"), indent=2)

    write_top_finds(ws)
    report_path = write_report(wb, total_new, categories)
    elapsed = time.time() - start_time
    sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}" if sheet_id else ""
    print(f"Run #{mem['runs']}: +{total_new} new, {max(0, ws.max_row-1)} total in {elapsed:.0f}s")
    if sheet_url:
        print(f"Master Sheet: {sheet_url}")
    write_google_doc(report_path, sheet_url)
    print(f"[{datetime.now(timezone.utc).isoformat()}] Done!")

if __name__ == "__main__":
    main()
